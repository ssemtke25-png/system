"""
탭14: 뉴스 클리핑
NAVER API HUB 뉴스 검색 API로 업무 관련 기사를 수집한다.

[동작]
- 키워드별로 최신순(sort=date) 검색
- pubDate 기준으로 지정 기간 내 기사만 필터링
- 중복 기사(제목 기준) 자동 제거
- 결과를 표로 표시 + 엑셀 다운로드

[secrets 설정]
NAVER_CLIENT_ID     = "..."
NAVER_CLIENT_SECRET = "..."
"""
import io
import re
import html
from datetime import datetime, timedelta, timezone
from urllib.parse import urlparse

import requests
import streamlit as st

API_URL = "https://naverapihub.apigw.ntruss.com/search/v1/news"
KST = timezone(timedelta(hours=9))

DEFAULT_KEYWORDS = [
    "지적재조사",
    "지적측량",
    "지적공부",
    "공간정보",
    "토지정보과",
    "토지관리",
    "개발부담금",
    "공인중개사",
]

# 제목·요약에 이 단어가 있으면 제외 (광고·홍보성 기사)
DEFAULT_EXCLUDE = [
    "학원", "수강", "인강", "특강", "개강", "강의",
    "합격", "자격증", "시험일정", "접수기간", "기출",
    "분양", "청약", "모델하우스", "입주자모집",
    "이벤트", "경품", "할인", "무료체험", "프로모션",
    "채용공고", "구인",
]

PERIOD_OPTIONS = {
    "최근 7일": 7,
    "최근 14일": 14,
    "최근 30일": 30,
    "최근 60일": 60,
    "최근 90일": 90,
}


# ── 유틸 ─────────────────────────────────────────────────────────────
def clean_text(s: str) -> str:
    """HTML 태그·엔티티 제거"""
    if not s:
        return ""
    s = re.sub(r"<[^>]+>", "", s)
    return html.unescape(s).strip()


def parse_pubdate(s: str):
    """'Thu, 11 Jun 2026 18:34:00 +0900' → datetime"""
    try:
        from email.utils import parsedate_to_datetime
        return parsedate_to_datetime(s)
    except Exception:
        return None


PRESS_ALIAS = {
    "yna": "연합뉴스", "yonhapnews": "연합뉴스",
    "chosun": "조선일보", "donga": "동아일보", "joongang": "중앙일보",
    "hani": "한겨레", "khan": "경향신문", "seoul": "서울신문",
    "mk": "매일경제", "hankyung": "한국경제", "sedaily": "서울경제",
    "imaeil": "매일신문", "yeongnam": "영남일보", "kyongbuk": "경북일보",
    "idaegu": "대구일보", "newsis": "뉴시스", "news1": "뉴스1",
    "edaily": "이데일리", "fnnews": "파이낸셜뉴스", "asiae": "아시아경제",
    "etnews": "전자신문", "dt": "디지털타임스", "zdnet": "지디넷",
    "kmib": "국민일보", "segye": "세계일보", "munhwa": "문화일보",
    "hankookilbo": "한국일보", "ohmynews": "오마이뉴스",
    "pressian": "프레시안", "nocutnews": "노컷뉴스",
    "gukjenews": "국제뉴스", "newspim": "뉴스핌", "ajunews": "아주경제",
}


def get_press(url: str) -> str:
    """URL 도메인에서 매체명 추출"""
    if not url:
        return ""
    try:
        host = urlparse(url).netloc.lower()
        # 서브도메인 제거 (news.naver.com, n.news.naver.com, m.yna.co.kr 등)
        host = re.sub(r"^(www|news|m|n|mobile|view|biz|land|realty)\.", "", host)

        parts = host.split(".")
        if not parts:
            return host

        # naver.com 계열은 '네이버뉴스'로
        if "naver" in parts:
            return "네이버뉴스"

        key = parts[0]
        # 남은 서브도메인 한 번 더 제거
        if key in ("news", "m", "n", "www") and len(parts) > 1:
            key = parts[1]

        return PRESS_ALIAS.get(key, key)
    except Exception:
        return ""


# ── 노이즈 필터 ──────────────────────────────────────────────────────
def apply_filters(rows, exclude_words, require_words):
    """제외 키워드 / 필수 키워드 적용

    exclude_words : 제목·요약에 있으면 제외
    require_words : 하나라도 있어야 통과 (비어 있으면 미적용)
    반환: (남은 기사, 제외된 건수)
    """
    if not exclude_words and not require_words:
        return rows, 0

    ex = [w.lower() for w in exclude_words if w]
    rq = [w.lower() for w in require_words if w]

    kept = []
    for r in rows:
        text = f"{r['제목']} {r['요약']}".lower()

        if ex and any(w in text for w in ex):
            continue
        if rq and not any(w in text for w in rq):
            continue

        kept.append(r)

    return kept, len(rows) - len(kept)


# ── API 호출 ─────────────────────────────────────────────────────────
def search_news(keyword: str, client_id: str, client_secret: str,
                days: int, max_items: int = 300):
    """키워드로 뉴스 검색 → 기간 내 기사만 반환"""
    headers = {
        "X-NCP-APIGW-API-KEY-ID": client_id,
        "X-NCP-APIGW-API-KEY": client_secret,
    }
    cutoff = datetime.now(KST) - timedelta(days=days)

    collected = []
    start = 1
    display = 100

    while start <= 1000 and len(collected) < max_items:
        try:
            resp = requests.get(
                API_URL,
                headers=headers,
                params={
                    "query": keyword,
                    "display": display,
                    "start": start,
                    "sort": "date",
                },
                timeout=10,
            )
        except Exception as e:
            return collected, f"요청 실패: {e}"

        if resp.status_code == 401:
            return collected, "인증 실패 — Client ID/Secret을 확인하세요."
        if resp.status_code != 200:
            try:
                msg = resp.json().get("errorMessage", resp.text[:100])
            except Exception:
                msg = resp.text[:100]
            return collected, f"[{resp.status_code}] {msg}"

        data = resp.json()
        items = data.get("items", [])
        if not items:
            break

        stop = False
        for it in items:
            dt = parse_pubdate(it.get("pubDate", ""))
            if dt and dt < cutoff:
                stop = True
                break
            collected.append({
                "날짜": dt.strftime("%Y-%m-%d") if dt else "",
                "시각": dt.strftime("%H:%M") if dt else "",
                "매체": get_press(it.get("originallink") or it.get("link")),
                "제목": clean_text(it.get("title")),
                "요약": clean_text(it.get("description")),
                "링크": it.get("originallink") or it.get("link", ""),
                "_dt": dt,
            })

        if stop or len(items) < display:
            break
        start += display

    return collected, None


def dedupe(rows):
    """제목 기준 중복 제거 (같은 기사 여러 매체 배포)"""
    seen = set()
    out = []
    for r in rows:
        key = re.sub(r"[^가-힣a-zA-Z0-9]", "", r["제목"])[:40]
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def to_excel(rows) -> bytes:
    """결과를 엑셀로 변환"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "뉴스클리핑"

    headers = ["번호", "날짜", "매체", "키워드", "제목", "요약", "링크"]
    widths  = [6, 12, 14, 14, 55, 70, 45]

    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(1, c, h)
        cell.font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    for i, r in enumerate(rows, start=1):
        row = i + 1
        vals = [i, r["날짜"], r["매체"], r.get("키워드", ""),
                r["제목"], r["요약"], r["링크"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row, c, v)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(c in (5, 6)),
                horizontal="center" if c in (1, 2, 3, 4) else "left",
            )
        # 링크 하이퍼링크
        if r["링크"]:
            lc = ws.cell(row, 7)
            lc.hyperlink = r["링크"]
            lc.font = Font(name="맑은 고딕", size=10,
                           color="0563C1", underline="single")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ─────────────────────────────────────────────────────
def render():
    st.caption("네이버 뉴스에서 업무 관련 기사를 키워드별로 수집합니다.")

    # 키 확인
    try:
        cid = st.secrets["NAVER_CLIENT_ID"]
        csec = st.secrets["NAVER_CLIENT_SECRET"]
    except Exception:
        st.error("❌ secrets에 NAVER_CLIENT_ID / NAVER_CLIENT_SECRET 이 없습니다.")
        st.code(
            'NAVER_CLIENT_ID = "발급받은_Client_ID"\n'
            'NAVER_CLIENT_SECRET = "발급받은_Client_Secret"',
            language="toml",
        )
        return

    col1, col2 = st.columns([2, 1])

    with col1:
        kw_text = st.text_area(
            "검색 키워드 (한 줄에 하나씩)",
            value="\n".join(DEFAULT_KEYWORDS),
            height=210,
            key="news_keywords",
        )
        keywords = [k.strip() for k in kw_text.split("\n") if k.strip()]

    with col2:
        period_label = st.selectbox("수집 기간", list(PERIOD_OPTIONS.keys()),
                                    index=2, key="news_period")
        days = PERIOD_OPTIONS[period_label]

        max_per_kw = st.number_input("키워드당 최대", 50, 1000, 200, step=50,
                                     key="news_max")
        do_dedupe = st.checkbox("중복 기사 제거", value=True, key="news_dedupe")
        st.caption(f"키워드 {len(keywords)}개 · {period_label}")

    # ── 노이즈 필터 ──────────────────────────────────
    with st.expander("🔧 노이즈 필터 (제외·필수 키워드)", expanded=False):
        st.caption(
            "제목이나 요약에 걸리는 단어로 기사를 걸러냅니다. "
            "쉼표(,) 또는 줄바꿈으로 구분하세요."
        )
        f1, f2 = st.columns(2)

        with f1:
            ex_text = st.text_area(
                "❌ 제외 — 이 단어가 있으면 뺍니다",
                value=", ".join(DEFAULT_EXCLUDE),
                height=150,
                key="news_exclude",
            )
        with f2:
            rq_text = st.text_area(
                "✅ 필수 — 이 중 하나는 있어야 합니다 (비우면 미적용)",
                value="",
                height=150,
                placeholder="예) 경북, 경상북도, 시청, 군청",
                key="news_require",
            )

        exclude_words = [w.strip() for w in re.split(r"[,\n]", ex_text) if w.strip()]
        require_words = [w.strip() for w in re.split(r"[,\n]", rq_text) if w.strip()]
        st.caption(f"제외 {len(exclude_words)}개 · 필수 {len(require_words)}개")

    if st.button("🔍 뉴스 수집", type="primary", key="btn_news"):
        if not keywords:
            st.warning("키워드를 하나 이상 입력하세요.")
            return

        all_rows = []
        errors = []
        prog = st.progress(0.0)
        status = st.empty()

        for i, kw in enumerate(keywords):
            status.text(f"검색 중… ({i+1}/{len(keywords)}) {kw}")
            rows, err = search_news(kw, cid, csec, days, max_per_kw)
            if err:
                errors.append({"키워드": kw, "오류": err})
            for r in rows:
                r["키워드"] = kw
            all_rows.extend(rows)
            prog.progress((i + 1) / len(keywords))

        prog.empty()
        status.empty()

        if errors:
            st.error("일부 키워드에서 오류가 발생했습니다.")
            st.dataframe(errors, use_container_width=True)

        if not all_rows:
            st.info("수집된 기사가 없습니다.")
            return

        raw_count = len(all_rows)

        # 노이즈 필터
        all_rows, filtered_out = apply_filters(
            all_rows, exclude_words, require_words
        )

        # 중복 제거
        before_dedupe = len(all_rows)
        if do_dedupe:
            all_rows = dedupe(all_rows)
        dup_removed = before_dedupe - len(all_rows)

        # 최신순 정렬
        all_rows.sort(key=lambda r: r["_dt"] or datetime.min.replace(tzinfo=KST),
                      reverse=True)

        st.session_state["news_rows"] = all_rows
        st.session_state["news_raw"] = raw_count
        st.session_state["news_filtered"] = filtered_out
        st.session_state["news_dup"] = dup_removed
        st.session_state["news_period_label"] = period_label

    # ── 결과 표시 ────────────────────────────────────
    rows = st.session_state.get("news_rows")
    if not rows:
        return

    raw_count    = st.session_state.get("news_raw", len(rows))
    filtered_out = st.session_state.get("news_filtered", 0)
    dup_removed  = st.session_state.get("news_dup", 0)

    parts = []
    if filtered_out:
        parts.append(f"노이즈 {filtered_out}건")
    if dup_removed:
        parts.append(f"중복 {dup_removed}건")

    if parts:
        st.success(
            f"✅ {len(rows)}건 "
            f"(원본 {raw_count}건 · {' · '.join(parts)} 제외)"
        )
    else:
        st.success(f"✅ {len(rows)}건 수집")

    if not rows:
        st.warning("남은 기사가 없습니다. 제외 키워드를 줄여보세요.")
        return

    # 키워드별 건수
    counts = {}
    for r in rows:
        counts[r["키워드"]] = counts.get(r["키워드"], 0) + 1
    if counts:
        cols = st.columns(min(len(counts), 4))
        for i, (k, v) in enumerate(sorted(counts.items(),
                                          key=lambda x: -x[1])):
            cols[i % len(cols)].metric(k, f"{v}건")

    st.markdown("---")

    # 필터
    fcol1, fcol2 = st.columns([1, 1])
    kw_filter = fcol1.multiselect(
        "키워드 필터", sorted(counts.keys()),
        default=sorted(counts.keys()), key="news_kwf"
    )
    presses = sorted({r["매체"] for r in rows if r["매체"]})
    press_filter = fcol2.multiselect(
        "매체 필터 (비우면 전체)", presses, default=[], key="news_pf"
    )

    view = [r for r in rows if r["키워드"] in kw_filter]
    if press_filter:
        view = [r for r in view if r["매체"] in press_filter]

    st.caption(f"표시 {len(view)}건")

    # 표 출력
    table = [
        {
            "날짜": r["날짜"],
            "매체": r["매체"],
            "키워드": r["키워드"],
            "제목": r["제목"],
            "링크": r["링크"],
        }
        for r in view
    ]
    st.dataframe(
        table,
        use_container_width=True,
        hide_index=True,
        column_config={
            "날짜":   st.column_config.TextColumn(width="small"),
            "매체":   st.column_config.TextColumn(width="small"),
            "키워드": st.column_config.TextColumn(width="small"),
            "제목":   st.column_config.TextColumn(width="large"),
            "링크":   st.column_config.LinkColumn("원문", display_text="열기",
                                                  width="small"),
        },
    )

    # 다운로드
    today = datetime.now(KST).strftime("%Y%m%d")
    st.download_button(
        "📥 엑셀 다운로드",
        data=to_excel(view),
        file_name=f"뉴스클리핑_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_news",
    )

    # 요약 보기
    with st.expander("📄 기사 요약 보기"):
        for r in view[:50]:
            st.markdown(
                f"**[{r['날짜']}] {r['제목']}**  \n"
                f"<span style='color:#888;font-size:0.85em'>"
                f"{r['매체']} · {r['키워드']}</span>  \n"
                f"{r['요약']}  \n"
                f"[원문 보기]({r['링크']})",
                unsafe_allow_html=True,
            )
            st.markdown("---")
        if len(view) > 50:
            st.caption(f"…외 {len(view)-50}건은 엑셀에서 확인하세요.")
