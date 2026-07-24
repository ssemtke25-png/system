"""
탭13: 개발부담금 부과·징수·물납 등 실적보고 취합
(서식1) 총괄표 + 시군별 파일 → 가나다순 세로 이어붙이기

[구조]
시트: '분기보고양식' 1개
행1~3: 헤더 (제목 + 2단 헤더)
행4부터: 데이터 (행 수 유동적)

[취합]
- 시군 파일의 행4부터 데이터를 가나다순으로 이어붙이기
- 총괄표의 기존 예시 데이터(○○시)는 삭제
- 데이터 없는 시군은 자동 건너뜀
"""
import io, re, zipfile, xml.etree.ElementTree as ET
import openpyxl
import streamlit as st
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from copy import copy

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

SHEET_CANDIDATES = ['분기보고양식', '분기보고', '실적보고']
HEADER_ROWS = 3          # 행1~3 헤더
DATA_START_ROW = 4       # 행4부터 데이터

# 경북 시군 가나다순
GYEONGBUK_ORDER = [
    "경산시","경주시","고령군","구미시","김천시","문경시","봉화군",
    "상주시","성주군","안동시","영덕군","영양군","영주시","영천시",
    "예천군","울릉군","울진군","의성군","청도군","청송군","칠곡군","포항시",
]


# ── 유틸 ─────────────────────────────────────────────────────────────
def shrink_styles(fb):
    fb.seek(0)
    try:
        zin = zipfile.ZipFile(fb, 'r')
        if 'xl/styles.xml' not in zin.namelist():
            fb.seek(0); return fb
        sb = zin.read('xl/styles.xml')
        ET.register_namespace('', NS_MAIN)
        root = ET.fromstring(sb)
        xfs = root.find(f'{{{NS_MAIN}}}cellStyleXfs')
        if xfs is None or len(xfs) <= 200:
            fb.seek(0); return fb
        first = xfs[0] if len(xfs) else None
        for x in list(xfs): xfs.remove(x)
        if first is not None: xfs.append(first)
        xfs.set('count','1')
        cxfs = root.find(f'{{{NS_MAIN}}}cellXfs')
        if cxfs:
            for x in cxfs:
                if x.get('xfId'): x.set('xfId','0')
        decl = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        nb = decl + ET.tostring(root, encoding='UTF-8')
        out = io.BytesIO()
        zout = zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED)
        for item in zin.namelist():
            zout.writestr(item, nb if item=='xl/styles.xml' else zin.read(item))
        zout.close(); zin.close(); out.seek(0); return out
    except Exception:
        fb.seek(0); return fb


def load_wb(fb, data_only=True):
    return openpyxl.load_workbook(shrink_styles(fb), data_only=data_only)


def find_sheet(wb, candidates=SHEET_CANDIDATES):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        for c in candidates:
            if c.replace(' ','') in name.replace(' ',''):
                return wb[name]
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def normalize_sigun(v):
    """시군명 정규화 (포항시 남구 → 포항시)"""
    if not v: return None
    s = str(v).strip()
    s = re.sub(r'^경상북도\s*', '', s)
    for name in GYEONGBUK_ORDER:
        base = re.sub(r'(시|군)$','',name)
        if base in s:
            return name
    return None


def sigun_sort_key(sigun_name):
    """가나다순 정렬 키"""
    if sigun_name in GYEONGBUK_ORDER:
        return GYEONGBUK_ORDER.index(sigun_name)
    return 999


def extract_sigun_from_filename(fname):
    m = re.match(r'^\d{1,3}[_.\s]+(.+?)\.', fname)
    raw = m.group(1) if m else fname
    raw = raw.replace('__','').replace('_',' ').strip()
    return normalize_sigun(raw)


def is_valid_row(ws, row, max_col):
    """실제 데이터가 있는 행인지"""
    return any(
        ws.cell(row, c).value not in (None, '', '○○시')
        for c in range(1, max_col+1)
    )


def is_example_row(ws, row):
    """예시 행(○○시)인지 판단"""
    v = ws.cell(row, 1).value
    return v and '○○' in str(v)


def clear_data_area(ws, start_row):
    """기존 데이터 영역 삭제"""
    if start_row > ws.max_row:
        return
    max_col = ws.max_column
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _xls_border_style(line_style):
    """xlrd 테두리 코드 → openpyxl 스타일명"""
    return {
        0: None, 1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
        5: "thick", 6: "double", 7: "hair", 8: "mediumDashed",
        9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot",
        12: "mediumDashDotDot", 13: "slantDashDot",
    }.get(line_style)


def _xls_halign(code):
    return {1: "left", 2: "center", 3: "right",
            4: "fill", 5: "justify", 6: "centerContinuous",
            7: "distributed"}.get(code)


def _xls_valign(code):
    return {0: "top", 1: "center", 2: "bottom",
            3: "justify", 4: "distributed"}.get(code)


def xls_to_xlsx(raw: bytes) -> bytes:
    """xls → xlsx 변환 (병합·글꼴·테두리·정렬·너비 모두 보존)"""
    import xlrd
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter as gcl

    book = xlrd.open_workbook(file_contents=raw, formatting_info=True)
    wb_new = openpyxl.Workbook()
    wb_new.remove(wb_new.active)

    for sname in book.sheet_names():
        ws_s = book.sheet_by_name(sname)
        ws_d = wb_new.create_sheet(sname)

        for r in range(ws_s.nrows):
            for c in range(ws_s.ncols):
                v = ws_s.cell_value(r, c)
                ctype = ws_s.cell_type(r, c)

                if ctype == xlrd.XL_CELL_DATE:
                    try:
                        from xlrd.xldate import xldate_as_datetime
                        v = xldate_as_datetime(v, book.datemode)
                    except Exception:
                        pass
                if v == '':
                    v = None

                cell = ws_d.cell(r + 1, c + 1, v)

                # ── 서식 복원 ──
                try:
                    xf = book.xf_list[ws_s.cell_xf_index(r, c)]
                    f = book.font_list[xf.font_index]

                    cell.font = Font(
                        name=f.name or "맑은 고딕",
                        size=(f.height / 20) if f.height else 11,
                        bold=bool(f.bold),
                        italic=bool(f.italic),
                        underline="single" if f.underline_type else None,
                    )

                    cell.alignment = Alignment(
                        horizontal=_xls_halign(xf.alignment.hor_align),
                        vertical=_xls_valign(xf.alignment.vert_align),
                        wrap_text=bool(xf.alignment.text_wrapped),
                    )

                    b = xf.border
                    sides = {}
                    for key, ls in (
                        ("top", b.top_line_style),
                        ("bottom", b.bottom_line_style),
                        ("left", b.left_line_style),
                        ("right", b.right_line_style),
                    ):
                        st_name = _xls_border_style(ls)
                        if st_name:
                            sides[key] = Side(style=st_name)
                    if sides:
                        cell.border = Border(**sides)

                    # 숫자 서식
                    fmt = book.format_map.get(xf.format_key)
                    if fmt and fmt.format_str and fmt.format_str != "General":
                        cell.number_format = fmt.format_str
                except Exception:
                    pass

        # 병합 셀
        for (rlo, rhi, clo, chi) in ws_s.merged_cells:
            try:
                ws_d.merge_cells(start_row=rlo + 1, end_row=rhi,
                                 start_column=clo + 1, end_column=chi)
            except Exception:
                pass

        # 열 너비
        try:
            for c in range(ws_s.ncols):
                info = ws_s.colinfo_map.get(c)
                if info and info.width:
                    ws_d.column_dimensions[gcl(c + 1)].width = info.width / 256
        except Exception:
            pass

        # 행 높이
        try:
            for r in range(ws_s.nrows):
                rh = ws_s.rowinfo_map.get(r)
                if rh and rh.height:
                    ws_d.row_dimensions[r + 1].height = rh.height / 20
        except Exception:
            pass

    buf = io.BytesIO()
    wb_new.save(buf)
    return buf.getvalue()


def copy_merges_from(src_ws, dst_ws, max_header_row=HEADER_ROWS):
    """헤더 영역의 병합 구조를 그대로 복사"""
    for mr in list(dst_ws.merged_cells.ranges):
        if mr.min_row <= max_header_row:
            try:
                dst_ws.unmerge_cells(str(mr))
            except Exception:
                pass
    for mr in src_ws.merged_cells.ranges:
        if mr.min_row <= max_header_row:
            try:
                dst_ws.merge_cells(str(mr))
            except Exception:
                pass


def autofit_columns(ws, max_col, start_row, end_row,
                    min_w=8, max_w=45, respect_existing=False):
    """열 너비 자동 조정 (### 방지) — 병합 셀 제외
    respect_existing=True 면 원본 서식의 너비를 유지한다.
    """
    from openpyxl.utils import get_column_letter as gcl

    merged_coords = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merged_coords.add((r, c))

    for c in range(1, max_col + 1):
        col_letter = gcl(c)
        existing = ws.column_dimensions[col_letter].width

        # 원본 너비가 있으면 그대로 둔다
        if respect_existing and existing:
            continue

        longest = 0
        for r in range(start_row, min(end_row + 1, ws.max_row + 1)):
            if (r, c) in merged_coords:
                continue
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            v = cell.value
            if v is None:
                continue
            fmt = cell.number_format or ""
            if "yy" in fmt.lower() or "mm" in fmt.lower():
                length = 12
            elif isinstance(v, (int, float)):
                length = len(f"{int(v):,}") + 2
            else:
                s = str(v)
                length = max(len(line) for line in s.split("\n"))
                han = sum(1 for ch in s if ord(ch) > 0x1100)
                length += han * 0.6
            longest = max(longest, length)

        if longest:
            ws.column_dimensions[col_letter].width = max(
                min_w, min(longest + 2, max_w)
            )


def copy_row_style(src_ws, src_row, dst_ws, dst_row, max_col):
    """행 스타일 복사"""
    for c in range(1, max_col + 1):
        src_cell = src_ws.cell(src_row, c)
        dst_cell = dst_ws.cell(dst_row, c)
        if isinstance(dst_cell, MergedCell):
            continue
        try:
            if src_cell.has_style:
                dst_cell.font          = copy(src_cell.font)
                dst_cell.border        = copy(src_cell.border)
                dst_cell.fill          = copy(src_cell.fill)
                dst_cell.number_format = copy(src_cell.number_format)
                dst_cell.alignment     = copy(src_cell.alignment)
        except Exception:
            pass


# ── 메인 취합 ────────────────────────────────────────────────────────
def aggregate(template_bytes, region_files):
    log, warns = [], []

    base_wb = load_wb(io.BytesIO(template_bytes), data_only=False)
    base_ws = find_sheet(base_wb)
    if not base_ws:
        raise ValueError("총괄표에서 '분기보고양식' 시트를 찾지 못했습니다.")

    max_col = base_ws.max_column

    # 시군 파일들의 최대 열 수 파악 (총괄표보다 넓을 수 있음)
    for fname, fbytes in region_files:
        try:
            probe_wb = load_wb(io.BytesIO(fbytes), data_only=True)
            probe_ws = find_sheet(probe_wb)
            if probe_ws:
                max_col = max(max_col, probe_ws.max_column)
        except Exception:
            pass

    # 스타일 템플릿용 행 저장 (예시 행 첫 번째)
    style_row = DATA_START_ROW

    # 데이터 영역을 지우기 전에 서식 템플릿을 미리 확보
    style_template = []
    for c in range(1, max_col + 1):
        cell = base_ws.cell(style_row, c)
        if isinstance(cell, MergedCell):
            style_template.append(None)
            continue
        style_template.append({
            "font":      copy(cell.font),
            "border":    copy(cell.border),
            "fill":      copy(cell.fill),
            "alignment": copy(cell.alignment),
            "fmt":       cell.number_format,
        })

    # 기존 데이터 영역 삭제 (값만)
    clear_data_area(base_ws, DATA_START_ROW)

    # ── 시군별 데이터 수집 ──────────────────────────────
    collected = []   # [(sigun, [row_values, ...]), ...]
    skipped = []

    for fname, fbytes in region_files:
        try:
            src_wb = load_wb(io.BytesIO(fbytes), data_only=True)
        except Exception as e:
            warns.append({"유형":"파일 오류","파일":fname,"설명":str(e)})
            continue

        src_ws = find_sheet(src_wb)
        if not src_ws:
            warns.append({"유형":"시트 없음","파일":fname,"설명":"분기보고양식 시트 없음"})
            continue

        src_max_col = max(src_ws.max_column, max_col)

        # 데이터 행 수집 (값 + 서식)
        rows = []
        for r in range(DATA_START_ROW, src_ws.max_row + 1):
            if is_example_row(src_ws, r):
                continue
            if not is_valid_row(src_ws, r, src_max_col):
                continue
            vals = []
            for c in range(1, max_col + 1):
                cell = src_ws.cell(r, c)
                v = cell.value
                if isinstance(v, str) and v.startswith('#'):
                    warns.append({"유형":"수식오류","파일":fname,
                                  "셀":f"{get_column_letter(c)}{r}","설명":f"'{v}' → 빈칸"})
                    v = None
                # 값 + 서식을 같이 저장
                vals.append({
                    "value": v,
                    "fmt": cell.number_format,
                })
            rows.append(vals)

        if not rows:
            skipped.append(fname)
            continue

        # 시군명 파악 (첫 행 A열 값)
        first_val = rows[0][0]["value"] if rows else None
        sigun = normalize_sigun(first_val) or extract_sigun_from_filename(fname)
        if not sigun:
            warns.append({"유형":"시군 인식 실패","파일":fname,"설명":"시군명 파악 불가"})
            sigun = "기타"

        collected.append((sigun, fname, rows))

    # ── 가나다순 정렬 ────────────────────────────────────
    collected.sort(key=lambda x: sigun_sort_key(x[0]))

    # ── 총괄표에 이어붙이기 ─────────────────────────────
    next_row = DATA_START_ROW
    for sigun, fname, rows in collected:
        for vals in rows:
            for c, item in enumerate(vals, start=1):
                cell = base_ws.cell(next_row, c)
                if isinstance(cell, MergedCell):
                    continue

                # 총괄표 원본 서식 적용 (테두리·글꼴·정렬)
                tpl = style_template[c - 1] if c - 1 < len(style_template) else None
                if tpl:
                    cell.font      = copy(tpl["font"])
                    cell.border    = copy(tpl["border"])
                    cell.fill      = copy(tpl["fill"])
                    cell.alignment = copy(tpl["alignment"])

                cell.value = item["value"]

                # 시군 원본의 숫자·날짜 서식이 있으면 우선
                if item["fmt"] and item["fmt"] != "General":
                    cell.number_format = item["fmt"]
                elif tpl and tpl["fmt"]:
                    cell.number_format = tpl["fmt"]

            base_ws.row_dimensions[next_row].height = \
                base_ws.row_dimensions[style_row].height
            next_row += 1
        log.append({"파일": fname, "시군": sigun, "행수": len(rows)})

    # ── 열 너비 자동 조정 (원본 너비 없는 열만) ──────────
    autofit_columns(base_ws, max_col, DATA_START_ROW, next_row - 1,
                    respect_existing=True)

    total_rows = next_row - DATA_START_ROW
    return base_wb, log, warns, collected, skipped, total_rows


# ── Streamlit UI ─────────────────────────────────────────────────────
def render():
    st.caption("시군별 개발부담금 실적보고 파일을 받아 총괄표(서식1)에 가나다순으로 이어붙입니다.")
    st.info(
        "📌 데이터가 없는 시군은 자동으로 건너뜁니다.\n\n"
        "📌 시군마다 행 수가 달라도 자동으로 처리됩니다.\n\n"
        "📌 가나다순(경산 → 경주 → ... → 포항)으로 정렬됩니다."
    )

    tpl_file = st.file_uploader(
        "① 총괄표(서식1) 파일 업로드",
        type=["xlsx","xls"],
        key="tpl_dev13"
    )
    region_files = st.file_uploader(
        "② 시군별 파일 업로드 (여러 개, xls/xlsx 모두 가능)",
        type=["xlsx","xls"],
        accept_multiple_files=True,
        key="region_dev13"
    )

    if tpl_file and region_files and st.button("🚀 취합 시작", key="btn_dev13"):

        def to_xlsx(fname, raw):
            """xls → xlsx 변환 (병합·서식 보존)"""
            if not fname.lower().endswith('.xls'):
                return raw
            try:
                return xls_to_xlsx(raw)
            except Exception as e:
                st.warning(f"⚠️ {fname} xls 변환 실패: {e}")
                return raw

        tpl_raw = to_xlsx(tpl_file.name, tpl_file.read())
        converted = [(f.name, to_xlsx(f.name, f.read())) for f in region_files]

        try:
            result_wb, log, warns, collected, skipped, total_rows = aggregate(
                tpl_raw, converted
            )

            out = io.BytesIO()
            result_wb.save(out)

            st.success(f"✅ 취합 완료! {len(collected)}개 시군 · 총 {total_rows}행")

            if skipped:
                st.info(f"📋 데이터 없어 건너뜀 ({len(skipped)}개): "
                        + ", ".join(skipped))

            st.download_button(
                "📥 취합 결과 다운로드",
                data=out.getvalue(),
                file_name="개발부담금_실적보고_취합.xlsx",
                key="dl_dev13"
            )

            col1, col2 = st.columns([1,1])
            with col1:
                st.markdown("**처리 결과 (가나다순)**")
                for item in log:
                    st.write(f"  ✅ {item['시군']} — {item['행수']}행")
            with col2:
                if warns:
                    st.warning(f"⚠️ 확인 필요 {len(warns)}건")
                    st.dataframe(warns, use_container_width=True)
                else:
                    st.info("특이사항 없이 정상 취합되었습니다.")

        except ValueError as e:
            st.error(f"❌ {e}")
        except Exception as e:
            st.error(f"오류: {e}")
            st.exception(e)
