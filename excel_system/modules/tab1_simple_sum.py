"""
탭1: 단순 합산
여러 엑셀 파일의 같은 위치 숫자를 모두 더한다. 수식이 있는 칸은 건드리지 않고 보존.
"""
import io
import openpyxl
import streamlit as st
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from modules.common import is_number, is_formula


def aggregate(file_bytes_list, file_names):
    warnings = []
    n_files = len(file_bytes_list)

    value_wbs = [openpyxl.load_workbook(b, data_only=True) for b in file_bytes_list]
    file_bytes_list[0].seek(0)
    base_wb = openpyxl.load_workbook(file_bytes_list[0], data_only=False)
    names = file_names

    all_sheets = []
    for vwb in value_wbs:
        for s in vwb.sheetnames:
            if s not in all_sheets:
                all_sheets.append(s)

    for sheet in all_sheets:
        present_idx = [i for i in range(n_files) if sheet in value_wbs[i].sheetnames]

        if len(present_idx) < n_files:
            missing_files = [names[i] for i in range(n_files) if i not in present_idx]
            warnings.append({
                "유형": "시트 누락", "시트": sheet, "셀": "-",
                "파일": ", ".join(missing_files),
                "설명": f"'{sheet}' 시트가 없어 해당 파일은 이 시트 합산에서 제외됨"
            })

        if not present_idx:
            continue

        if sheet not in base_wb.sheetnames:
            base_wb.create_sheet(sheet)

        # 무한 로딩 방지 안전장치
        max_r = min(max(value_wbs[i][sheet].max_row for i in present_idx), 300)
        max_c = min(max(value_wbs[i][sheet].max_column for i in present_idx), 100)
        base_ws = base_wb[sheet]

        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell_addr = f"{get_column_letter(c)}{r}"
                base_cell = base_ws.cell(r, c)

                if isinstance(base_cell, MergedCell):
                    continue
                if is_formula(base_cell.value):
                    continue

                total = 0
                any_number = False
                text_files = []

                for i in present_idx:
                    vws = value_wbs[i][sheet]
                    if r > vws.max_row or c > vws.max_column:
                        continue
                    v = vws.cell(r, c).value
                    if is_number(v):
                        total += v
                        any_number = True
                    elif v is not None and isinstance(v, str):
                        text_files.append((names[i], v))

                if any_number and text_files:
                    warnings.append({
                        "유형": "텍스트 혼입", "시트": sheet, "셀": cell_addr,
                        "파일": ", ".join(fn for fn, _ in text_files),
                        "설명": "숫자가 들어가야 할 칸에 텍스트 발견 ("
                                + ", ".join(f"{fn}='{val}'" for fn, val in text_files)
                                + ") → 0으로 처리하여 합산함"
                    })

                if any_number:
                    base_cell.value = total

    return base_wb, warnings


def render():
    """탭1 화면을 그린다. app.py에서 with tab1: render() 형태로 호출."""
    st.caption("여러 파일의 같은 위치 숫자를 모두 더합니다. 수식이 있는 칸은 건드리지 않고 그대로 보존합니다.")
    files1 = st.file_uploader("파일 업로드", type=["xlsx"], accept_multiple_files=True, key="up1")

    if files1 and st.button("🚀 취합 시작", key="btn1"):
        try:
            file_bytes_list = [io.BytesIO(f.read()) for f in files1]
            file_names = [f.name for f in files1]

            result_wb, warns = aggregate(file_bytes_list, file_names)

            o = io.BytesIO()
            result_wb.save(o)

            st.success("취합이 완료되었습니다.")
            st.download_button("📥 다운로드", o.getvalue(), "result.xlsx", key="dl1")

            if warns:
                st.warning(f"⚠️ 확인이 필요한 항목 {len(warns)}건이 발견되었습니다. (결과는 정상 생성됨)")
                st.dataframe(warns, use_container_width=True)
            else:
                st.info("특이사항 없이 정상적으로 합산되었습니다.")

        except Exception as e:
            st.error(f"오류: {e}")
            st.exception(e)
