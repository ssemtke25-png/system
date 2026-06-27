"""
탭2: 총괄표 채우기 (시군구)
시군구별로 작성된 여러 파일을 하나의 '총괄표' 서식에 자동으로 채워 넣는다.
파일명(NN_지역명_...)에서 지역을 인식하고, 시트가 '지역이 행에 나열'인지
'지역이 열에 나열'인지 자동 판별해서 자기 지역 자리만 채운다.
수식이 있는 칸은 절대 건드리지 않는다.
"""
import io
import openpyxl
import streamlit as st
from openpyxl.cell.cell import MergedCell

from modules.common import (
    is_formula, get_safe_value, region_key, is_valid_region,
    extract_own_region_from_filename, get_sheet_by_index, target_keys_for_region,
)


def detect_layout(ws, max_scan_row=30, max_scan_col=30):
    region_like_count_in_col_a = 0
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        v = ws.cell(r, 1).value
        if is_valid_region(v):
            region_like_count_in_col_a += 1
    if region_like_count_in_col_a >= 3:
        return 'row'

    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        hits = sum(
            1 for c in range(1, min(ws.max_column, max_scan_col) + 1)
            if is_valid_region(ws.cell(r, c).value)
        )
        if hits >= 3:
            return 'col'
    return None


def find_region_start_row(ws, target_keys, max_scan_row=120):
    max_row = min(ws.max_row, max_scan_row)
    for r in range(1, max_row + 1):
        label = ws.cell(r, 1).value
        key = region_key(label) if label else None
        if key in target_keys:
            return r
    return None


def get_base_group_size(base_ws, max_scan_row=120):
    max_row = min(base_ws.max_row, max_scan_row)
    region_rows = []
    for r in range(1, max_row + 1):
        label = base_ws.cell(r, 1).value
        if is_valid_region(label):
            region_rows.append(r)
    if len(region_rows) >= 2:
        return region_rows[1] - region_rows[0]
    return 1


def find_header_row(ws, target_keys=None, max_scan_row=30):
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        if target_keys:
            hits = sum(
                1 for c in range(1, ws.max_column + 1)
                if region_key(ws.cell(r, c).value) in target_keys
            )
            if hits >= 1:
                return r
        else:
            hits = sum(
                1 for c in range(1, ws.max_column + 1)
                if is_valid_region(ws.cell(r, c).value)
            )
            if hits >= 3:
                return r
    return None


def find_region_col_in_sheet(ws, header_row, target_keys):
    if header_row is None:
        return {}
    col_for_key = {}
    for c in range(1, ws.max_column + 1):
        label = ws.cell(header_row, c).value
        key = region_key(label) if label else None
        if key in target_keys:
            col_for_key.setdefault(key, c)
    return col_for_key


def fill_row_layout(base_ws, src_ws, own_key, warnings, sheet_title):
    target_keys = target_keys_for_region(own_key)
    if not target_keys:
        return 0

    max_col = base_ws.max_column
    group_size = get_base_group_size(base_ws)
    count = 0

    for key in target_keys:
        base_start = find_region_start_row(base_ws, {key})
        src_start = find_region_start_row(src_ws, {key})

        if base_start is None or src_start is None:
            continue

        for offset in range(group_size):
            gr_base = base_start + offset
            gr_src = src_start + offset
            for c in range(2, max_col + 1):
                base_cell = base_ws.cell(gr_base, c)
                if isinstance(base_cell, MergedCell):
                    continue
                if is_formula(base_cell.value):
                    continue

                src_val = src_ws.cell(gr_src, c).value
                src_val = get_safe_value(src_val)

                if src_val is None:
                    continue

                base_cell.value = src_val
        count += 1

    return count


def fill_col_layout(base_ws, src_ws, own_key, warnings, sheet_title):
    target_keys = target_keys_for_region(own_key)
    if not target_keys:
        return 0

    base_header_row = find_header_row(base_ws)
    src_header_row = find_header_row(src_ws, target_keys=set(target_keys))

    if base_header_row is None or src_header_row is None:
        return 0

    base_col_for_key = find_region_col_in_sheet(base_ws, base_header_row, set(target_keys))
    src_col_for_key = find_region_col_in_sheet(src_ws, src_header_row, set(target_keys))

    max_row = min(base_ws.max_row, 120)
    count = 0
    for key in target_keys:
        base_col = base_col_for_key.get(key)
        src_col = src_col_for_key.get(key)
        if base_col is None or src_col is None:
            continue

        row_offset = src_header_row - base_header_row
        for r in range(base_header_row + 1, max_row + 1):
            base_cell = base_ws.cell(r, base_col)
            if isinstance(base_cell, MergedCell):
                continue
            if is_formula(base_cell.value):
                continue

            src_r = r + row_offset
            src_val = src_ws.cell(src_r, src_col).value if src_r <= src_ws.max_row else None
            src_val = get_safe_value(src_val)

            if src_val is None:
                continue

            base_cell.value = src_val
        count += 1
    return count


def is_standard_structure(src_ws, max_scan_row=120):
    max_row = min(src_ws.max_row, max_scan_row)
    region_row_count = 0
    for r in range(1, max_row + 1):
        if is_valid_region(src_ws.cell(r, 1).value):
            region_row_count += 1
    return region_row_count >= 5


def check_structure_mismatch(src_ws, sheet_title, fname, warnings):
    if not is_standard_structure(src_ws):
        warnings.append({
            "유형": "⚠ 양식 임의변경 감지", "시트": sheet_title, "셀": "-",
            "파일": fname,
            "설명": "이 파일은 자기 지역만 남기고 삭제되는 등 편집이 감지되었습니다. (데이터는 제자리에 정상 취합됩니다)"
        })


def fill_master_template(template_bytes, region_files_with_names):
    log = []
    warnings = []

    base_wb = openpyxl.load_workbook(template_bytes, data_only=False)
    sheet_count = len(base_wb.sheetnames)

    layouts = {}
    for idx in range(sheet_count):
        ws = get_sheet_by_index(base_wb, idx)
        layouts[idx] = detect_layout(ws)

    for fname, fbytes in region_files_with_names:
        own_key = extract_own_region_from_filename(fname)
        if not own_key:
            warnings.append({
                "유형": "지역 인식 실패", "시트": "-", "셀": "-",
                "설명": f"'{fname}' 파일명에서 지역명을 추출할 수 없어 건너뜀"
            })
            continue

        fbytes.seek(0)
        wb_values = openpyxl.load_workbook(fbytes, data_only=True)

        if len(wb_values.sheetnames) != sheet_count:
            warnings.append({
                "유형": "시트 개수 다름", "시트": "-", "셀": "-",
                "설명": f"'{fname}'은 시트 {len(wb_values.sheetnames)}개 (총괄표는 {sheet_count}개)"
            })

        first_sheet = get_sheet_by_index(wb_values, 0)
        if first_sheet is not None and layouts.get(0) == 'row':
            check_structure_mismatch(first_sheet, "전체 파일", fname, warnings)

        for idx in range(min(sheet_count, len(wb_values.sheetnames))):
            base_ws = get_sheet_by_index(base_wb, idx)
            src_ws = get_sheet_by_index(wb_values, idx)
            layout = layouts.get(idx)

            if layout == 'row':
                n = fill_row_layout(base_ws, src_ws, own_key, warnings, base_ws.title)
            elif layout == 'col':
                n = fill_col_layout(base_ws, src_ws, own_key, warnings, base_ws.title)
            else:
                n = 0
            if n:
                log.append({"파일": fname, "시트": base_ws.title, "방식": layout, "처리건수": n})

    return base_wb, log, warnings


def render():
    """탭2 화면을 그린다. app.py에서 with tab2: render() 형태로 호출."""
    st.caption("시군구별로 작성된 여러 파일을 하나의 '총괄표' 서식에 자동으로 채워 넣습니다.")
    st.info("📌 파일명은 'NN_지역명_...' 형식이어야 합니다 (예: 01_포항시_...).")

    template_file = st.file_uploader("① 총괄표(서식) 파일 업로드", type=["xlsx"], key="template_up")
    region_files = st.file_uploader("② 시군구별 파일 업로드 (여러 개)", type=["xlsx"], accept_multiple_files=True, key="region_up")

    if template_file and region_files and st.button("🚀 총괄표 채우기 시작", key="btn2"):
        try:
            template_bytes = io.BytesIO(template_file.read())
            region_files_with_names = [(f.name, io.BytesIO(f.read())) for f in region_files]

            result_wb, log, warns = fill_master_template(template_bytes, region_files_with_names)

            o = io.BytesIO()
            result_wb.save(o)

            st.success("총괄표 채우기가 완료되었습니다.")
            st.download_button("📥 다운로드", o.getvalue(), "총괄표_결과.xlsx", key="dl2")

            if warns:
                st.warning(f"⚠️ 확인이 필요한 항목 {len(warns)}건이 발견되었습니다. (결과는 정상 생성됨)")
                st.dataframe(warns, use_container_width=True)
            else:
                st.info("특이사항 없이 정상적으로 채워졌습니다.")

            with st.expander("처리 로그 보기 (성공 내역)"):
                st.dataframe(log, use_container_width=True)

        except Exception as e:
            st.error(f"오류: {e}")
            st.exception(e)
