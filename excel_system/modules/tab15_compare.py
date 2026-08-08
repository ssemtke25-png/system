# -*- coding: utf-8 -*-
"""
탭15: 분기·연차 비교 시각화
이미 취합된 결과 파일 여러 개를 업로드 → 같은 지표를 기간별로 비교.

[설계 원리]
- 기간 라벨은 사용자가 자유 지정 (분기/연차 무관)
- 각 파일 → (기간,시군,지표,값) long-format → pivot
  ⇒ 파일이 2개든 12개든 코드 한 벌로 처리
- 시군명 정규화로 '포항시 남구'='포항남' 매칭 (표준목록 방식)

[입력] 취합 완료본 xlsx 여러 개 (탭2~14에서 나온 결과)
[출력] 지표별 비교표 + 추이 차트 + 엑셀 다운로드
"""
import io
import re
from datetime import datetime

import pandas as pd
import openpyxl
import streamlit as st


# ══════════════════════════════════════════════
# 시군명 정규화 (매칭 키)
# ══════════════════════════════════════════════
SIGUN_CANON = ['포항남구', '포항북구', '경주', '김천', '안동', '구미', '영주', '영천',
               '상주', '문경', '경산', '군위', '의성', '청송', '영양', '영덕', '청도',
               '고령', '성주', '칠곡', '예천', '봉화', '울진', '울릉']

# 표기 변형 → 표준키 (실무에서 발견되는 변형을 여기 추가)
SIGUN_ALIAS = {
    '포항시남구': '포항남구', '포항남': '포항남구', '포항시 남구': '포항남구',
    '포항시북구': '포항북구', '포항북': '포항북구', '포항시 북구': '포항북구',
}


def norm_sigun(raw):
    """'포항시 남구','포항남','포항 남구' → '포항남구' 로 통일.
    단순 문자제거는 '청송→송' 오류가 나므로 표준목록 매칭 방식 사용."""
    if not raw:
        return None
    s = re.sub(r'[\s\u3000]', '', str(raw))
    if s in SIGUN_ALIAS:
        return SIGUN_ALIAS[s]
    s2 = re.sub(r'(시|군)$', '', s)
    if s2 in SIGUN_ALIAS:
        return SIGUN_ALIAS[s2]
    for canon in SIGUN_CANON:
        if s2 == canon or s2 == canon.replace('구', ''):
            return canon
    return s2 or s


def num(v):
    return v if isinstance(v, (int, float)) else 0


# ══════════════════════════════════════════════
# 파일 → long-format 추출
# ══════════════════════════════════════════════
# 지표 정의: (시트명, 추출함수) — 취합 서식이 바뀌면 여기만 수정
def _extract_jongsaja(ws, period, rows):
    for r in range(5, 28):
        nm = ws.cell(r, 1).value
        if nm and nm != '계':
            sg = norm_sigun(nm)
            rows.append((period, sg, '소속공인중개사', num(ws.cell(r, 4).value)))
            rows.append((period, sg, '중개보조원', num(ws.cell(r, 5).value)))


def _extract_gaeeop(ws, period, rows):
    """각 시군은 4행 블록(중개사/중개인/법인/계). 시군명은 첫 행에만.
    '계' 행을 앵커로 찾아 총수(C)·신규(D)·폐업(E)·등록취소(F)·이전(G)을 읽고
    순증감 = 신규 - 폐업 - 등록취소 + 이전 을 산출.
    (단순 4행 고정점프는 빈 행/서식변형에 취약하므로 '계' 탐색 방식)"""
    r = 3
    maxr = ws.max_row
    while r <= maxr:
        nm = ws.cell(r, 1).value
        if nm and str(nm).strip() not in ('합계', '구분', ''):
            sg = norm_sigun(nm)
            gyr = None
            for k in range(5):                       # 앵커 아래 5행 내 '계' 탐색
                if str(ws.cell(r + k, 2).value).strip() == '계':
                    gyr = r + k
                    break
            if gyr:
                total = num(ws.cell(gyr, 3).value)
                singyu = num(ws.cell(gyr, 4).value)   # 신규등록
                pyeeop = num(ws.cell(gyr, 5).value)   # 폐업
                chwiso = num(ws.cell(gyr, 6).value)   # 등록취소
                ijeon = num(ws.cell(gyr, 7).value)    # 이전
                net = singyu - pyeeop - chwiso + ijeon
                rows.append((period, sg, '개업공인중개사', total))
                rows.append((period, sg, '신규등록', singyu))
                rows.append((period, sg, '폐업', pyeeop))
                rows.append((period, sg, '순증감', net))
                r = gyr + 1
                continue
        r += 1


def _extract_dansok(ws, period, rows):
    for r in range(7, 30):
        nm = ws.cell(r, 1).value
        if nm and nm != '계':
            rows.append((period, norm_sigun(nm), '단속업소수', num(ws.cell(r, 2).value)))


def _extract_viol(ws, period, rows, label):
    """위반유형 시트는 [근거조문|위반내용|계|시군별...] 구조.
    '계' 총계 행의 C열(3) 값이 곧 전체 건수 → 그 값을 직접 읽는다.
    (조문×시군 전체합 방식은 소계·중복행이 섞이면 이중합산되어 취약)"""
    tot = None
    for r in range(1, min(ws.max_row, 12) + 1):
        a = str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else ''
        b = str(ws.cell(r, 2).value).strip() if ws.cell(r, 2).value else ''
        # 헤더(구분/근거조문) 다음의 '계' 총계행 찾기
        if a == '계' or b == '계':
            tot = num(ws.cell(r, 3).value)
            break
    if tot is None:                                  # '계' 라벨 없으면 헤더 다음행이 총계
        for r in range(4, min(ws.max_row, 12) + 1):
            v = ws.cell(r, 3).value
            if isinstance(v, (int, float)):
                tot = v
                break
    rows.append((period, '(전체)', f'위반-{label}', num(tot)))


SHEET_EXTRACTORS = [
    ('2.시도별 중개업 종사자수(분기)', _extract_jongsaja),
    ('3.개업공인중개사 증감내역', _extract_gaeeop),
    ('5.개업공인중개사 지도단속실적', _extract_dansok),
    ('6.고발조치', lambda ws, p, rows: _extract_viol(ws, p, rows, '고발조치')),
    ('7.등록취소', lambda ws, p, rows: _extract_viol(ws, p, rows, '등록취소')),
    ('8.업무정지', lambda ws, p, rows: _extract_viol(ws, p, rows, '업무정지')),
    ('9.과태료처분', lambda ws, p, rows: _extract_viol(ws, p, rows, '과태료')),
    ('10.자격취소, 자격정지(끝)', lambda ws, p, rows: _extract_viol(ws, p, rows, '자격취소·정지')),
]


def extract_long(file_bytes, period):
    """한 파일(바이트) → long-format rows [(기간, 시군, 지표, 값), ...]"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = []
    for sheet_name, fn in SHEET_EXTRACTORS:
        if sheet_name in wb.sheetnames:
            try:
                fn(wb[sheet_name], period, rows)
            except Exception as e:
                st.warning(f"[{period}] '{sheet_name}' 추출 중 오류: {e}")
    return rows


# ══════════════════════════════════════════════
# 자동 감지 모드 (등록 안 된 임의 취합본용)
# ══════════════════════════════════════════════
# 지표명으로 잡히면 합계열이라 기본 제외할 헤더
SKIP_HEADERS = {'계', '합계', '소계', '총계', '구분', '시군', '시도',
                '연번', '순번', '번호', '비고'}


def is_sigun_strict(raw):
    """자동 감지 전용: 표준목록에 정확히 매칭될 때만 True.
    (norm_sigun은 매칭 실패 시 입력을 그대로 반환하므로 감지에는 부적합 —
     '중개업 종사자수' 같은 제목까지 시군으로 오인함)"""
    if not raw:
        return False
    s = re.sub(r'[\s\u3000]', '', str(raw))
    if s in SIGUN_ALIAS:
        return True
    s2 = re.sub(r'(시|군)$', '', s)
    if s2 in SIGUN_ALIAS:
        return True
    for canon in SIGUN_CANON:
        if s2 == canon or s2 == canon.replace('구', ''):
            return True
    return False


def autodetect_sheet(ws):
    """시트에서 시군이 세로축인지 판별하고, 지표 열들을 자동 감지.
    반환 {sig_col, rows, header_r, metrics:[{col,name,is_total}]} 또는 None."""
    best_col, best_rows = None, []
    for c in range(1, min(ws.max_column, 6) + 1):
        hits = [r for r in range(1, min(ws.max_row, 80) + 1)
                if is_sigun_strict(ws.cell(r, c).value)]
        if len(hits) > len(best_rows):
            best_col, best_rows = c, hits
    if len(best_rows) < 5:              # 시군이 세로로 5개 미만이면 시군축 아님
        return None
    sig_col = best_col
    r0 = min(best_rows)

    header_r = None                     # 헤더행: 데이터 위쪽 문자열 2개↑ 첫 행
    for r in range(r0 - 1, max(0, r0 - 6), -1):
        txt = sum(1 for c in range(sig_col, ws.max_column + 1)
                  if isinstance(ws.cell(r, c).value, str))
        if txt >= 2:
            header_r = r
            break

    metrics = []
    for c in range(sig_col + 1, ws.max_column + 1):
        numeric = sum(1 for r in best_rows
                      if isinstance(ws.cell(r, c).value, (int, float)))
        if numeric < len(best_rows) * 0.5:
            continue
        h = ws.cell(header_r, c).value if header_r else None
        name = str(h).strip().replace('\n', ' ') if h else f'열{c}'
        metrics.append({'col': c, 'name': name, 'is_total': name in SKIP_HEADERS})

    # 헤더 영역 텍스트 (데이터 시작 위 6행) — Gemini 해석용 재료
    # 숫자가 대부분인 행(데이터·계 행)은 제외: 헤더 텍스트만 남긴다.
    head_top = max(1, r0 - 6)
    header_grid = []
    for r in range(head_top, r0):
        row_cells = []
        n_num = n_txt = 0
        for c in range(sig_col, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                n_num += 1
            elif isinstance(v, str) and v.strip():
                n_txt += 1
            row_cells.append('' if v is None else str(v).replace('\n', ' ').strip())
        if n_num > n_txt:               # 숫자가 더 많으면 데이터행 → 스킵
            continue
        header_grid.append((r, row_cells))
    # 첫 데이터행 샘플(값 예시) — 어떤 열이 숫자인지 Gemini가 보게
    sample = []
    for c in range(sig_col, ws.max_column + 1):
        v = ws.cell(r0, c).value
        sample.append('' if v is None else str(v))

    return {'sig_col': sig_col, 'rows': best_rows, 'header_r': header_r,
            'metrics': metrics, 'header_grid': header_grid,
            'sample_row': sample, 'col_start': sig_col}


def scan_workbook(file_bytes):
    """파일 전체를 스캔 → {시트명: 감지결과}. UI에서 지표 확인·수정용."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = {}
    for sn in wb.sheetnames:
        det = autodetect_sheet(wb[sn])
        if det and det['metrics']:
            out[sn] = det
    return out


# ══════════════════════════════════════════════
# Gemini 헤더 해석 (복잡한 2단·기호 헤더에 이름 붙이기)
# 원칙: 숫자는 절대 넘기지 않는다. 헤더 텍스트만 보내 '이름표'만 받는다.
# ══════════════════════════════════════════════
def _gemini_model():
    """tab7과 동일하게 secrets의 GEMINI_API_KEY 사용."""
    import google.generativeai as genai
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
    return genai.GenerativeModel("gemini-2.5-flash-lite")


def interpret_headers_gemini(sheet_name, det):
    """헤더 영역 텍스트를 Gemini에 보내 각 지표 열의 사람이 읽을 이름을 받는다.
    반환 {col: 해석된_이름}. 실패 시 {} (호출부에서 기존 name 유지)."""
    import json

    cols = [m['col'] for m in det['metrics']]
    col_start = det['col_start']

    # 헤더 그리드를 사람이 보듯 텍스트로 (열 위치 표기)
    lines = []
    for r, cells in det['header_grid']:
        tagged = []
        for i, v in enumerate(cells):
            c = col_start + i
            if v:
                tagged.append(f"[C{c}]{v}")
        if tagged:
            lines.append("  " + " ".join(tagged))
    header_text = "\n".join(lines) if lines else "(헤더 텍스트 없음)"

    # 대상 열 목록
    targets = ", ".join(f"C{m['col']}(현재:{m['name']})" for m in det['metrics'])

    prompt = f"""다음은 한국 행정 통계표의 헤더 영역이다. 여러 행에 걸쳐 병합된 다단 헤더일 수 있다.
시트명: {sheet_name}

[헤더 영역 — 각 셀 앞 [C숫자]는 열 위치]
{header_text}

[이름을 붙일 대상 열]
{targets}

각 대상 열이 나타내는 지표의 이름을, 위 헤더를 세로로 종합해서 사람이 이해할 한국어 명사구로 정하라.
규칙:
- 기호((A=C+K), (E) 등)나 숫자만 있는 열은 상위 헤더를 참고해 의미 있는 이름으로.
- 상위 대분류가 있으면 '대분류-소분류' 형태 (예: '행정조치-등록취소').
- 합계/소계 성격의 열은 이름 끝에 '(계)'를 붙여라.
- 판단 불가하면 빈 문자열.
반드시 아래 JSON만 출력 (설명 금지):
{{"C3": "이름", "C5": "이름"}}"""

    try:
        model = _gemini_model()
        resp = model.generate_content(prompt)
        txt = resp.text.strip()
        txt = re.sub(r'^```(?:json)?|```$', '', txt, flags=re.MULTILINE).strip()
        data = json.loads(txt)
        result = {}
        for m in det['metrics']:
            key = f"C{m['col']}"
            if key in data and isinstance(data[key], str) and data[key].strip():
                result[m['col']] = data[key].strip()
        return result
    except Exception as e:
        st.warning(f"[{sheet_name}] 헤더 AI 해석 실패 (수동 이름 사용): {e}")
        return {}


def extract_auto(file_bytes, period, plan):
    """자동 감지 계획(plan)에 따라 long-format 추출.
    plan = {시트명: {col: 표시할_지표명}} — 사람이 화면에서 고르고 이름 붙인 것.
    지표명 앞에 시트태그를 붙여 서로 다른 시트의 동명 지표 충돌을 막는다."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = []
    for sn, colmap in plan.items():
        if sn not in wb.sheetnames or not colmap:
            continue
        ws = wb[sn]
        det = autodetect_sheet(ws)
        if not det:
            continue
        sig_col = det['sig_col']
        for r in det['rows']:
            raw = ws.cell(r, sig_col).value
            if not is_sigun_strict(raw):     # det['rows']는 이미 걸러졌으나 이중안전
                continue
            sg = norm_sigun(raw)
            for c, label in colmap.items():
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)):
                    rows.append((period, sg, label, v))
    return rows


def list_sheets_with_sigun(file_bytes):
    """파일 안에서 시군 세로축이 감지되는 시트만 골라 [(시트명, 감지결과), ...] 반환.
    '한 파일에 여러 달(4·5·6월)이 시트로 나뉜' 경우를 비교하기 위한 목록."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = []
    for sn in wb.sheetnames:
        det = autodetect_sheet(wb[sn])
        if det and det['metrics']:
            out.append((sn, det))
    return out


def extract_sheet_as_period(file_bytes, sheet_name, period, colmap):
    """한 파일의 특정 시트를 하나의 기간으로 추출.
    colmap = {col: 지표명}. 여러 시트를 서로 다른 period로 부르면 기간 비교가 된다."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = []
    if sheet_name not in wb.sheetnames:
        return rows
    ws = wb[sheet_name]
    det = autodetect_sheet(ws)
    if not det:
        return rows
    sig_col = det['sig_col']
    for r in det['rows']:
        raw = ws.cell(r, sig_col).value
        if not is_sigun_strict(raw):
            continue
        sg = norm_sigun(raw)
        for c, label in colmap.items():
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                rows.append((period, sg, label, v))
    return rows


def guess_period_from_sheet(sheet_name):
    """시트명에서 기간 라벨 추론. '월별내역(26.4월...)' → '26.4월' 등."""
    m = re.search(r"(\d{2,4})\s*[.\-년]\s*(\d{1,2})\s*월", sheet_name)
    if m:
        yy, mm = m.group(1), int(m.group(2))
        return f"{yy}.{mm}월"
    m = re.search(r'([1-4])\s*(?:/4)?\s*분기', sheet_name)
    if m:
        return f"{m.group(1)}Q"
    # 괄호 안 내용
    m = re.search(r'[\(（]([^\)）]{1,12})[\)）]', sheet_name)
    if m:
        return m.group(1)[:10]
    return sheet_name[:12]



    name = filename.rsplit('.', 1)[0]
    # 2026_2분기, 2026년 2분기, 2026-2Q 등
    m = re.search(r'(20\d{2}).*?([1-4])\s*(?:/4)?\s*분기', name)
    if m:
        return f"{m.group(1)} {m.group(2)}Q"
    m = re.search(r'(20\d{2})[\s_\-]*[qQ]\s*([1-4])', name)
    if m:
        return f"{m.group(1)} {m.group(2)}Q"
    m = re.search(r'(20\d{2})\s*년', name)
    if m:
        return f"{m.group(1)}년"
    return name[:20]


# ══════════════════════════════════════════════
# 비교표 생성 + 엑셀 출력
# ══════════════════════════════════════════════
def generate_comment_gemini(df, chosen):
    """완성된 비교표(pivot)를 텍스트로 요약해 Gemini에 보내 해석 코멘트를 받는다.
    숫자는 코드가 확정한 값만 전달 — Gemini는 읽고 문장만 쓴다.
    반환 코멘트 문자열, 실패 시 ''."""
    periods = sorted(df['기간'].unique())
    if len(periods) < 2:
        return ""

    # 지표별 기간 추이를 텍스트 표로 (전체 합계 기준)
    lines = []
    for ind in chosen:
        sub = df[df['지표'] == ind]
        vals = []
        for p in periods:
            v = sub[sub['기간'] == p]['값'].sum()
            vals.append(f"{p}={v:,.0f}")
        lines.append(f"- {ind}: " + ", ".join(vals))
    table_text = "\n".join(lines)

    prompt = f"""다음은 경상북도 시군별 행정 통계의 기간별 비교 결과다.
숫자는 이미 확정된 값이며, 너는 이 수치를 근거로 추세 해석 코멘트만 작성한다.
숫자를 새로 계산하거나 바꾸지 마라.

[기간별 지표 추이]
{table_text}

작성 규칙:
- 3~5문장의 간결한 행정보고체 (개조식 아님, '~함/~임' 체).
- 증가/감소 추세와 그 의미를 짚되, 데이터에 없는 사실은 지어내지 마라.
- 여러 지표가 있으면 지표 간 관계(동반 증감 등)를 언급하면 좋다.
- 과장·단정 금지. '~로 보임', '~로 판단됨' 등 신중한 표현.
- 표제어 없이 본문만 출력."""

    try:
        model = _gemini_model()
        resp = model.generate_content(prompt)
        return resp.text.strip()
    except Exception as e:
        st.warning(f"해석 코멘트 생성 실패: {e}")
        return ""



def build_cover(ws, df, title="분기 비교 현황", comment=""):
    """요약보고 표지 시트 작성. 모든 수치를 df에서 자동 계산 (Gemini 불필요).
    최신기간 기준 현황 + 전기간 대비 증감률. comment가 있으면 하단에 AI 코멘트."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    NAVY, BLUE, RED, GREY = "1F3864", "2E74B5", "C00000", "595959"
    HEADFILL = "DEEBF7"
    F = "맑은 고딕"
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    periods = sorted(df['기간'].unique())
    cur = periods[-1]
    prev = periods[0] if len(periods) >= 2 else None

    def gv(period, ind):
        s = df[(df['기간'] == period) & (df['지표'] == ind)]
        return int(s['값'].sum())

    def top_n(period, ind, n=5):
        s = df[(df['기간'] == period) & (df['지표'] == ind) & (df['시군'] != '(전체)')]
        return s.groupby('시군')['값'].sum().sort_values(ascending=False).head(n)

    def fmt(n):
        return f"{n:,}"

    # 수치 수집
    sosok = gv(cur, '소속공인중개사')
    bojo = gv(cur, '중개보조원')
    gaeeop = gv(cur, '개업공인중개사')
    singyu = gv(cur, '신규등록')
    pyeeop = gv(cur, '폐업')
    net = gv(cur, '순증감')
    dansok = gv(cur, '단속업소수')
    viol_labels = ['고발조치', '등록취소', '업무정지', '과태료', '자격취소·정지']
    viols = {lab: gv(cur, '위반-' + lab) for lab in viol_labels}
    viol_total = sum(viols.values())

    # 상위 문구
    sosok_top = top_n(cur, '소속공인중개사')
    sosok_txt = ", ".join(f"{sg} {int(v)}명" for sg, v in sosok_top.items())
    gaeeop_top = top_n(cur, '개업공인중개사')
    gaeeop_txt = ", ".join(f"{sg} {v/gaeeop*100:.1f}%" for sg, v in gaeeop_top.items()) if gaeeop else ""
    dansok_top = top_n(cur, '단속업소수')
    dansok_txt = ", ".join(f"{sg} {int(v)}개소" for sg, v in dansok_top.items())

    # 열너비
    for col, w in [('A', 2.5), ('B', 15), ('C', 13), ('G', 12), ('H', 11), ('I', 2.5)]:
        ws.column_dimensions[col].width = w

    def put(coord, val, size=11, bold=False, color="000000",
            name=F, ha=None, va="bottom", wrap=False, fill=None):
        c = ws[coord]
        c.value = val
        c.font = Font(name=name, size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        return c

    # 우상단 발행 정보
    ws.merge_cells('G1:H1'); put('G1', '토지정보과', 9, False, GREY, ha="center", va="center")
    ws.merge_cells('G2:H2')
    put('G2', datetime.now().strftime('%Y. %m.  .(  )'), 9, False, GREY, ha="center", va="center")

    # 제목
    ws.merge_cells('B3:H4')
    put('B3', title, 22, True, NAVY, ha="center", va="center")

    # 목적 박스
    ws.merge_cells('B6:H7')
    put('B6', "◈ 개업공인중개사 등록·종사자 현황 및 지도단속 실적 관리를 통해\n"
              "    건전한 부동산 중개질서 확립 및 소비자 보호에 기여",
        12, True, NAVY, ha="left", va="center", wrap=True)

    # □ 종사자 현황
    put('B9', "□ 종사자 현황", 13, True, NAVY)
    ws.merge_cells('B10:C10'); put('B10', "○ 소속공인중개사 ", 11.5, True, "000000")
    ws.merge_cells('D10:H10')
    put('D10', f"{fmt(sosok)}명 · 중개보조원 {fmt(bojo)}명", 11.5, True, BLUE)
    ws.merge_cells('B11:H11')
    put('B11', f"  • 소속중개사 상위 : {sosok_txt} 등", 11)
    ws.merge_cells('B12:H12')
    put('B12', f"  • 개업공인중개사 총 {fmt(gaeeop)}명", 11.5, True, "000000")
    ws.merge_cells('B13:H13')
    put('B13', f"    - 시군별 비중 : {gaeeop_txt} 등", 11)

    # □ 개업 증감 현황
    put('B15', "□ 개업공인중개사 증감 현황", 13, True, NAVY)
    trend = "감소" if net < 0 else ("증가" if net > 0 else "보합")
    ws.merge_cells('B16:H16')
    put('B16', f"○ 순증감 : {net:+d}명 ({trend})", 11.5, True,
        RED if net < 0 else NAVY)
    ws.merge_cells('B17:H17')
    put('B17', f"  • 신규등록 {singyu}명, 폐업 {pyeeop}명", 11.5, True, "000000")
    ws.merge_cells('B18:H18')
    reason = ("폐업이 신규등록을 상회하여 개업 수 감소 추세" if net < 0
              else "신규등록이 폐업을 상회하여 개업 수 증가 추세" if net > 0
              else "신규등록과 폐업이 균형")
    put('B18', f"    - {reason}", 11)

    # □ 지도단속 현황
    put('B20', "□ 지도단속 현황", 13, True, NAVY)
    ws.merge_cells('B21:H21')
    put('B21', f"○ 단속 업소 {fmt(dansok)}개소 · 행정조치 {viol_total}건", 11.5, True, "000000")
    ws.merge_cells('B22:H22')
    put('B22', f"  • 단속 상위 : {dansok_txt} 등", 11)
    ws.merge_cells('B23:H23')
    put('B23', "  • 위반유형별 행정조치", 11)

    # 위반유형 표 (r24 헤더, r25 건수)
    hdr = ['구분'] + viol_labels
    for j, h in enumerate(hdr):
        cc = ws.cell(24, 3 + j, h.replace('자격취소·정지', '자격취소\n정지'))
        cc.font = Font(name=F, size=9.5, bold=True, color=NAVY)
        cc.fill = PatternFill("solid", fgColor=HEADFILL)
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = box
    ws.cell(25, 3, "건수").font = Font(name=F, size=10, bold=True)
    ws.cell(25, 3).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(25, 3).border = box
    for j, lab in enumerate(viol_labels):
        cc = ws.cell(25, 4 + j, viols[lab])
        cc.font = Font(name=F, size=10)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = box
    ws.row_dimensions[24].height = 27.75

    # □ 증감 현황 (전기간 대비)
    if prev:
        put('B27', f"□ 증감 현황 ({prev} 대비)", 13, True, NAVY)
        def rate(ind):
            c, p = gv(cur, ind), gv(prev, ind)
            return (c - p) / p * 100 if p else 0.0
        parts = [f"소속중개사 {rate('소속공인중개사'):+.1f}%",
                 f"중개보조원 {rate('중개보조원'):+.1f}%",
                 f"단속 {rate('단속업소수'):+.1f}%"]
        ws.merge_cells('B28:H28')
        put('B28', "○ " + " · ".join(parts), 11, True, "000000")

    # AI 해석 코멘트 (있으면 증감현황 아래)
    if comment:
        put('B31', "□ 분석 의견 (AI 초안)", 13, True, NAVY)
        ws.merge_cells('B32:H37')
        cc = ws['B32']
        cc.value = comment
        cc.font = Font(name=F, size=10.5, color="000000")
        cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for r in range(32, 38):
            ws.row_dimensions[r].height = 15

    # 하단 발행처
    put('H49', "토지정보과", 11, False, "000000", ha="center")

    # 표지는 눈금선 숨김
    ws.sheet_view.showGridLines = False


def build_generic_cover(ws, df, chosen, title="기간 비교 현황",
                        comment="", unit_hint=""):
    """범용 표지 — 지표명이 무엇이든 데이터에서 자동 생성.
    공인중개사 전용 build_cover와 달리 업무별 하드코딩 문구가 없어
    개발부담금 등 어떤 취합본에도 붙는다."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    NAVY, RED, GREY, GREEN = "1F3864", "C00000", "595959", "1E8449"
    F = "맑은 고딕"
    thin = Side(style="thin", color="BFBFBF")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    periods = sorted(df['기간'].unique())
    cur, first = periods[-1], periods[0]

    def gv(period, ind):
        s = df[(df['기간'] == period) & (df['지표'] == ind)]
        return s['값'].sum()

    def top_n(period, ind, n=5):
        s = df[(df['기간'] == period) & (df['지표'] == ind) & (df['시군'] != '(전체)')]
        return s.groupby('시군')['값'].sum().sort_values(ascending=False).head(n)

    def fmt(v):
        return f"{v:,.0f}"

    for col, w in [('A', 2.5), ('B', 16), ('C', 15), ('D', 15), ('E', 15),
                   ('F', 15), ('G', 13), ('H', 11)]:
        ws.column_dimensions[col].width = w

    def put(coord, val, size=11, bold=False, color="000000", ha=None,
            va="center", wrap=False, fill=None):
        c = ws[coord]
        c.value = val
        c.font = Font(name=F, size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
        if fill:
            c.fill = PatternFill("solid", fgColor=fill)
        return c

    ws.merge_cells('F1:H1'); put('F1', '토지정보과', 9, False, GREY, ha="center")
    ws.merge_cells('F2:H2'); put('F2', datetime.now().strftime('%Y. %m.  .(  )'),
                                  9, False, GREY, ha="center")
    ws.merge_cells('B3:H4'); put('B3', title, 22, True, NAVY, ha="center")

    ws.merge_cells('B6:H6')
    put('B6', f"◈ 비교기간 : {first} ~ {cur}  (총 {len(periods)}개 기간)"
              + (f"   ·   단위 : {unit_hint}" if unit_hint else ""),
        11.5, True, NAVY, ha="left")

    # □ 기간별 총계 추이표
    put('B8', "□ 기간별 총계 추이", 13, True, NAVY)
    hr = 9
    put(f'B{hr}', "지표", 10.5, True, "FFFFFF", ha="center", fill=NAVY).border = box
    for j, p in enumerate(periods):
        cc = ws.cell(hr, 3 + j, p)
        cc.font = Font(name=F, size=10.5, bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor=NAVY)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = box
    jcol = 3 + len(periods)
    cc = ws.cell(hr, jcol, "증감률")
    cc.font = Font(name=F, size=10.5, bold=True, color="FFFFFF")
    cc.fill = PatternFill("solid", fgColor=NAVY)
    cc.alignment = Alignment(horizontal="center", vertical="center"); cc.border = box

    for i, ind in enumerate(chosen):
        rr = hr + 1 + i
        put(f'B{rr}', ind, 10, False, "000000", ha="left").border = box
        for j, p in enumerate(periods):
            cc = ws.cell(rr, 3 + j, gv(p, ind))
            cc.font = Font(name=F, size=10); cc.number_format = "#,##0"
            cc.alignment = Alignment(horizontal="right", vertical="center")
            cc.border = box
        v0, v1 = gv(first, ind), gv(cur, ind)
        rate = (v1 - v0) / v0 * 100 if v0 else 0.0
        cc = ws.cell(rr, jcol, rate / 100)
        cc.number_format = "+0.0%;-0.0%;0.0%"
        cc.font = Font(name=F, size=10, bold=True,
                       color=GREEN if rate > 0 else (RED if rate < 0 else GREY))
        cc.alignment = Alignment(horizontal="center", vertical="center"); cc.border = box

    # □ 최신기간 상위 시군 (첫 지표 기준)
    base = hr + 2 + len(chosen)
    put(f'B{base}', f"□ 최신기간({cur}) 상위 시군", 13, True, NAVY)
    lead = chosen[0]
    tops = top_n(cur, lead)
    tot = gv(cur, lead)
    ws.merge_cells(f'B{base+1}:H{base+1}')
    txt = ", ".join(f"{sg} {fmt(v)}" + (f"({v/tot*100:.1f}%)" if tot else "")
                    for sg, v in tops.items())
    put(f'B{base+1}', f"○ {lead} : {txt} 등", 10.5, False, "000000",
        ha="left", wrap=True)

    # □ 증감 요약
    base2 = base + 3
    put(f'B{base2}', f"□ 증감 요약 ({first} 대비)", 13, True, NAVY)
    parts = []
    for ind in chosen:
        d = gv(cur, ind) - gv(first, ind)
        parts.append(f"{ind} {'+' if d >= 0 else ''}{fmt(d)}")
    ws.merge_cells(f'B{base2+1}:H{base2+1}')
    put(f'B{base2+1}', "○ " + " · ".join(parts), 10.5, False, "000000",
        ha="left", wrap=True)

    # □ AI 해석 의견
    if comment:
        base3 = base2 + 3
        put(f'B{base3}', "□ 분석 의견 (AI 초안)", 13, True, NAVY)
        ws.merge_cells(f'B{base3+1}:H{base3+7}')
        cc = ws[f'B{base3+1}']
        cc.value = comment
        cc.font = Font(name=F, size=10.5)
        cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.sheet_view.showGridLines = False


def build_pivot(df, indicator):
    """지표 하나 → pivot (행=시군/유형, 열=기간, +증감/증감률)"""
    sub = df[df['지표'] == indicator]
    if sub.empty:
        return None
    pv = sub.pivot_table(index='시군', columns='기간', values='값',
                         aggfunc='sum').fillna(0)
    pv = pv[sorted(pv.columns)]       # 기간 시간순
    cols = list(pv.columns)
    if len(cols) >= 2:
        pv['증감'] = pv[cols[-1]] - pv[cols[0]]
        pv['증감률(%)'] = pv.apply(
            lambda row: round((row[cols[-1]] - row[cols[0]]) / row[cols[0]] * 100, 1)
            if row[cols[0]] else 0.0, axis=1)
        pv = pv.sort_values(cols[-1], ascending=False)
    return pv


def to_excel(pivots, df=None, cover_title="분기 비교 현황", comment="",
             generic=False, chosen=None):
    """지표별 pivot들 → 시각화 엑셀 (기존 양식과 동일 톤).
    df + generic=False → 공인중개사 전용 표지(build_cover).
    df + generic=True  → 범용 표지(build_generic_cover, 지표명 무관).
    comment가 있으면 표지 하단(또는 별도 시트)에 AI 해석 코멘트를 넣는다."""
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    NAVY, BLUE, GREY, WHITE = "1A2B44", "2E6FD0", "5B6D85", "FFFFFF"
    LINE, GREEN, RED = "D6DEEA", "1E8449", "C0392B"
    F = "맑은 고딕"
    thin = Side(style="thin", color=LINE)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 표지 — 맨 앞
    if df is not None and not df.empty:
        cover = wb.create_sheet("요약보고")
        try:
            if generic:
                cols = chosen or list(dict.fromkeys(df['지표'].tolist()))
                build_generic_cover(cover, df, cols, title=cover_title,
                                    comment=comment, unit_hint="")
            else:
                build_cover(cover, df, title=cover_title, comment=comment)
        except Exception as e:
            cover["B2"] = f"표지 생성 오류: {e}"
    elif comment:
        # 표지 없이 코멘트만
        cs = wb.create_sheet("AI 해석")
        cs["B2"] = "AI 해석 코멘트"
        cs["B2"].font = Font(name=F, size=14, bold=True, color=NAVY)
        cs["B4"] = comment
        cs["B4"].font = Font(name=F, size=11)
        cs["B4"].alignment = Alignment(wrap_text=True, vertical="top")
        cs.merge_cells("B4:H20")
        cs.column_dimensions["A"].width = 2.5
        for col in "BCDEFGH":
            cs.column_dimensions[col].width = 14
        cs.sheet_view.showGridLines = False

    for indicator, pv in pivots.items():
        if pv is None:
            continue
        safe = re.sub(r'[\\/*?:\[\]]', '', indicator)[:28]
        ws = wb.create_sheet(safe)
        ws["A1"] = f"{indicator} 기간 비교"
        ws["A1"].font = Font(name=F, size=15, bold=True, color=NAVY)

        # 헤더
        hrow = 3
        headers = ['구분'] + [str(c) for c in pv.columns]
        for j, h in enumerate(headers):
            c = ws.cell(hrow, 1 + j, h)
            c.font = Font(name=F, size=10.5, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
        # 데이터
        period_cols = [c for c in pv.columns if c not in ('증감', '증감률(%)')]
        for i, (idx, row) in enumerate(pv.iterrows()):
            rr = hrow + 1 + i
            cc = ws.cell(rr, 1, idx)
            cc.font = Font(name=F, size=10); cc.border = border
            cc.alignment = Alignment(horizontal="left", vertical="center")
            for j, col in enumerate(pv.columns):
                v = row[col]
                cell = ws.cell(rr, 2 + j, round(v, 1) if isinstance(v, float) else v)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col == '증감률(%)':
                    delta = row.get('증감', 0)
                    color = GREEN if delta > 0 else (RED if delta < 0 else GREY)
                    cell.font = Font(name=F, size=10, bold=True, color=color)
                    cell.number_format = "+0.0;-0.0;0.0"
                elif col == '증감':
                    color = GREEN if v > 0 else (RED if v < 0 else GREY)
                    cell.font = Font(name=F, size=10, bold=True, color=color)
                else:
                    cell.font = Font(name=F, size=10)
        lastrow = hrow + len(pv)
        ws.column_dimensions["A"].width = 14
        for c in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12

        # 추이 차트 (기간 컬럼만)
        if len(period_cols) >= 2 and len(pv) <= 30:
            ch = BarChart(); ch.type = "col"
            ch.title = f"{indicator} 기간 추이"
            first = 2
            last = 1 + len(period_cols)
            data = Reference(ws, min_col=first, max_col=last,
                             min_row=hrow, max_row=lastrow)
            cats = Reference(ws, min_col=1, min_row=hrow + 1, max_row=lastrow)
            ch.add_data(data, titles_from_data=True)
            ch.set_categories(cats)
            ch.height = 10; ch.width = 24; ch.gapWidth = 60
            ws.add_chart(ch, f"A{lastrow + 3}")

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ══════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════
def render():
    st.caption("취합 완료본 여러 개를 올리면 기간별로 비교합니다. (분기·연차 무관)")

    mode = st.radio(
        "분석 모드",
        ["공인중개사 (정밀)", "자동 감지 (아무 취합본)", "단일 파일 시트별 비교"],
        horizontal=True, key="cmp_mode",
    )
    auto = mode.startswith("자동")
    single = mode.startswith("단일")

    if auto:
        st.caption("🔎 시군이 세로로 나열된 시트를 자동으로 찾아 지표를 감지합니다. "
                   "감지된 지표명이 이상하면 직접 고쳐 쓸 수 있어요.")
    elif single:
        st.caption("📑 한 파일 안에 여러 기간(예: 4·5·6월)이 시트로 나뉜 경우, "
                   "각 시트를 기간으로 잡아 비교합니다. 파일 1개만 올리세요.")

    files = st.file_uploader(
        "취합 결과 파일 업로드 (여러 개)",
        type=["xlsx"],
        accept_multiple_files=not single,   # 단일 모드는 파일 1개만
        key="cmp_files",
    )
    if not files:
        if single:
            st.info("한 파일 안에 여러 기간(월·분기)이 시트로 나뉜 취합본 1개를 올려주세요.")
        else:
            st.info("비교할 취합본을 2개 이상 올려주세요. 예: 2025_3분기.xlsx, 2026_1분기.xlsx")
        return

    # ══ 단일 파일 시트별 비교 모드 ══
    if single:
        f = files if not isinstance(files, list) else files[0]
        fbytes = f.getvalue()
        sheets = list_sheets_with_sigun(fbytes)
        if len(sheets) < 2:
            st.error("시군이 세로로 된 시트가 2개 이상 있어야 비교할 수 있습니다. "
                     f"(감지된 시트: {len(sheets)}개)")
            return

        st.markdown("**① 비교할 시트(기간) 선택**")
        sheet_names = [sn for sn, _ in sheets]
        picked = st.multiselect(
            "시트", sheet_names, default=sheet_names,
            key="cmp_single_sheets",
        )
        if len(picked) < 2:
            st.info("비교할 시트를 2개 이상 선택하세요.")
            return

        # 각 시트에 기간 라벨
        st.markdown("**② 각 시트의 기간 라벨 확인·수정**")
        sheet_periods = {}
        for sn in picked:
            c1, c2 = st.columns([3, 2])
            c1.text(f"📑 {sn}")
            sheet_periods[sn] = c2.text_input(
                "기간", value=guess_period_from_sheet(sn),
                key=f"cmp_sp_{sn}", label_visibility="collapsed",
            )

        # 지표 감지·선택 (첫 시트 기준, AI 해석 옵션 포함)
        st.markdown("**③ 감지된 지표 확인·선택** (첫 시트 기준, 모든 시트 동일 적용)")
        det0 = dict(sheets)[picked[0]]
        use_ai = st.checkbox(
            "🤖 AI로 지표명 자동 해석 (복잡한 2단·기호 헤더)",
            value=False, key="cmp_single_ai",
        )
        ai_names = st.session_state.get("cmp_single_ai_names", {})
        if use_ai and st.button("🤖 AI 해석 실행", key="cmp_single_ai_run"):
            with st.spinner("헤더를 AI가 해석하는 중..."):
                ai_names = interpret_headers_gemini(picked[0], det0)
                st.session_state["cmp_single_ai_names"] = ai_names

        colmap = {}
        for m in det0['metrics']:
            c1, c2 = st.columns([1, 3])
            use = c1.checkbox("사용", value=not m['is_total'],
                              key=f"cmp_sm_use_{m['col']}",
                              label_visibility="collapsed")
            default_name = ai_names.get(m['col'], m['name'])
            label = c2.text_input("지표명", value=default_name,
                                  key=f"cmp_sm_lab_{m['col']}",
                                  label_visibility="collapsed")
            if use and label.strip():
                colmap[m['col']] = label.strip()

        if st.button("📊 비교 생성", type="primary", key="cmp_single_run"):
            if not colmap:
                st.error("지표를 하나 이상 선택하세요.")
                return
            all_rows = []
            for sn in picked:
                all_rows += extract_sheet_as_period(
                    fbytes, sn, sheet_periods[sn].strip(), colmap)
            if not all_rows:
                st.error("데이터를 추출하지 못했습니다.")
                return
            df = pd.DataFrame(all_rows, columns=['기간', '시군', '지표', '값'])
            st.session_state["cmp_df"] = df
            st.session_state["cmp_is_auto"] = True   # 표지는 공인중개사 전용이므로 생략

        _render_results()   # 결과 표시·다운로드 (공통)
        return

    # 각 파일에 기간 라벨 지정
    st.markdown("**① 각 파일의 기간 라벨 확인·수정**")
    periods = {}
    for i, f in enumerate(files):
        col1, col2 = st.columns([3, 2])
        col1.text(f"📄 {f.name}")
        default = guess_period(f.name)
        periods[i] = col2.text_input(
            "기간 라벨", value=default, key=f"cmp_period_{i}",
            label_visibility="collapsed",
        )

    if len({p.strip() for p in periods.values()}) < len(periods):
        st.warning("⚠ 기간 라벨이 중복됩니다. 서로 다르게 입력하세요.")

    # ── 자동 감지 모드: 첫 파일 스캔 → 지표 확인·이름수정 ──
    plan = None
    if auto:
        st.markdown("**② 감지된 지표 확인·선택** (첫 파일 기준, 모든 파일에 동일 적용)")
        scan = scan_workbook(files[0].getvalue())
        if not scan:
            st.error("시군이 세로로 나열된 데이터 시트를 찾지 못했습니다. "
                     "시군명이 한 열에 세로로 있는 취합본인지 확인하세요.")
            return

        use_ai = st.checkbox(
            "🤖 AI로 지표명 자동 해석 (복잡한 2단·기호 헤더 이름붙이기)",
            value=False, key="cmp_use_ai",
            help="병합된 다단 헤더나 (A), (E) 같은 기호 헤더를 AI가 사람이 읽을 이름으로 바꿔 "
                 "기본값에 채웁니다. 숫자는 AI에 보내지 않으며, 결과는 직접 수정 가능합니다.",
        )
        ai_names = {}                    # {시트명: {col: AI이름}}
        if use_ai and st.button("🤖 AI 해석 실행", key="cmp_ai_run"):
            with st.spinner("헤더를 AI가 해석하는 중..."):
                for sn, det in scan.items():
                    ai_names[sn] = interpret_headers_gemini(sn, det)
            st.session_state["cmp_ai_names"] = ai_names
        ai_names = st.session_state.get("cmp_ai_names", {})

        plan = {}
        for sn, det in scan.items():
            sheet_ai = ai_names.get(sn, {})
            with st.expander(f"📄 {sn}  (시군 {len(det['rows'])}개 · 지표 {len(det['metrics'])}개)",
                             expanded=True):
                colmap = {}
                for m in det['metrics']:
                    c1, c2 = st.columns([1, 3])
                    use = c1.checkbox(
                        "사용", value=not m['is_total'],
                        key=f"cmp_use_{sn}_{m['col']}",
                        label_visibility="collapsed",
                    )
                    # AI 이름이 있으면 기본값으로 사용 (사람이 다시 수정 가능)
                    default_name = sheet_ai.get(m['col'], m['name'])
                    label = c2.text_input(
                        "지표명", value=default_name,
                        key=f"cmp_lab_{sn}_{m['col']}",
                        label_visibility="collapsed",
                    )
                    if use and label.strip():
                        colmap[m['col']] = label.strip()
                if colmap:
                    plan[sn] = colmap

    btn_label = "📊 비교 생성"
    if st.button(btn_label, type="primary", key="cmp_run"):
        all_rows = []
        prog = st.progress(0.0)
        for i, f in enumerate(files):
            data = f.getvalue()
            if auto:
                all_rows += extract_auto(data, periods[i].strip(), plan)
            else:
                all_rows += extract_long(data, periods[i].strip())
            prog.progress((i + 1) / len(files))
        prog.empty()

        if not all_rows:
            st.error("데이터를 추출하지 못했습니다. 지표 선택 또는 서식을 확인하세요.")
            return

        df = pd.DataFrame(all_rows, columns=['기간', '시군', '지표', '값'])
        st.session_state["cmp_df"] = df
        st.session_state["cmp_is_auto"] = auto

    _render_results()


def _render_results():
    """비교 결과 표시·엑셀 다운로드 (모든 모드 공통).
    st.session_state['cmp_df']와 ['cmp_is_auto']를 읽어 렌더링."""
    df = st.session_state.get("cmp_df")
    if df is None:
        return
    is_auto = st.session_state.get("cmp_is_auto", False)

    indicators = list(df['지표'].unique())
    st.markdown("**④ 비교할 지표 선택**")
    chosen = st.multiselect(
        "지표", indicators, default=indicators[:min(4, len(indicators))],
        key="cmp_ind",
    )
    if not chosen:
        st.info("지표를 하나 이상 선택하세요.")
        return

    pivots = {}
    for ind in chosen:
        pv = build_pivot(df, ind)
        pivots[ind] = pv
        if pv is not None:
            st.markdown(f"#### {ind}")
            st.dataframe(pv, use_container_width=True)

    # 표지: 정밀 모드 → 전용 표지 / 자동·단일 모드 → 범용 표지
    st.markdown("**⑤ 표지 제목**")
    cover_title = st.text_input(
        "표지 제목", value="분기 비교 현황" if not is_auto else "기간 비교 현황",
        key="cmp_title", label_visibility="collapsed",
    )
    cover_df = df

    # 🤖 AI 해석 코멘트 (완성된 비교표를 읽고 추세 문장 생성)
    st.markdown("**⑥ AI 해석 코멘트** (선택)")
    comment = st.session_state.get("cmp_comment", "")
    if st.button("🤖 해석 코멘트 생성", key="cmp_comment_run"):
        with st.spinner("비교 결과를 AI가 해석하는 중..."):
            comment = generate_comment_gemini(df, chosen)
            st.session_state["cmp_comment"] = comment
    if comment:
        st.info(comment)
        st.caption("※ 위 코멘트는 확정된 수치를 근거로 AI가 작성한 초안입니다. "
                   "보고 전 사실관계를 확인하세요.")

    # 엑셀 다운로드 (자동·단일 모드는 범용 표지)
    xlsx = to_excel(
        {k: v for k, v in pivots.items() if v is not None},
        df=cover_df, cover_title=cover_title, comment=comment,
        generic=is_auto, chosen=chosen,
    )
    today = datetime.now().strftime("%Y%m%d")
    st.download_button(
        "📥 비교 엑셀 다운로드 (차트 포함)",
        data=xlsx,
        file_name=f"분기비교_{today}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", key="cmp_dl",
    )
