"""
탭5(지적재조사 조정금), 탭6(의견접수·이의신청)이 공유하는 핵심 로직.

두 탭은 서식 구조(22개 지역별 시트 + '총괄' 시트, A열에 '연도' 또는 '합계'
라벨로 행을 구분하는 방식)가 완전히 동일해서 같은 함수를 함께 쓴다.
업로드된 시군구 파일에서 연도별 데이터를 찾아, 총괄표 서식의 해당 지역
시트 중 같은 연도 행에 그대로 옮겨 넣는다.

'총괄' 또는 '총괄(자동입력)'처럼 이름이 '총괄'로 시작하는 시트는
이번 취합과 무관한 별도 집계용 시트이므로, 서식 파일 쪽이든 업로드된
시군구 파일 쪽이든 절대 읽거나 쓰지 않는다.
"""
import openpyxl
from openpyxl.cell.cell import MergedCell

from modules.common import is_formula, get_safe_value
from modules.tab3_realestate_monthly import extract_own_region_re, sort_key_for_filename_re

JIJEOK_SHEET_MAP = {
    '포항남': '포항 남구', '포항북': '포항 북구', '경주': '경주', '김천': '김천', '안동': '안동',
    '구미': '구미', '영주': '영주', '영천': '영천', '상주': '상주', '문경': '문경',
    '경산': '경산', '군위': '군위', '의성': '의성', '청송': '청송', '영양': '영양',
    '영덕': '영덕', '청도': '청도', '고령': '고령', '성주': '성주', '칠곡': '칠곡',
    '예천': '예천', '봉화': '봉화', '울진': '울진', '울릉': '울릉'
}


def target_keys_for_region(own_key):
    """포항은 남구/북구 2개 시트 모두 대상으로 처리해야 함."""
    if own_key == '포항':
        return ['포항남', '포항북']
    return [own_key] if own_key else []


def is_protected_summary_sheet(sheet_name):
    """'총괄', '총괄(자동입력)'처럼 '총괄'로 시작하는 이름의 시트는
    이번 취합 대상이 아니므로 절대 읽거나 쓰지 않는다."""
    if not sheet_name:
        return False
    return sheet_name.strip().replace(" ", "").startswith("총괄")


def fill_jijeok_template(template_bytes, region_files_with_names):
    """
    template_bytes: 지역별 시트가 있는 총괄표 서식 (xlsx)
    region_files_with_names: [(파일명, BytesIO), ...] 시군구별 파일들

    동작:
    - 파일명에서 지역을 인식해 그 지역의 시트를 찾는다 (포항은 남/북 둘 다).
    - '총괄' 계열 시트는 서식이든 원본 파일이든 절대 건드리지 않는다.
    - 원본 파일의 A열에서 '합계'/'OOOO년' 라벨이 있는 행을 찾아, 서식의
      같은 라벨이 있는 행에 B열부터 그대로 옮긴다. 수식·병합 칸은 보존한다.
    """
    log = []
    warnings = []

    base_wb = openpyxl.load_workbook(template_bytes, data_only=False)
    sorted_files = sorted(region_files_with_names, key=lambda x: sort_key_for_filename_re(x[0]))

    for fname, fbytes in sorted_files:
        own_key = extract_own_region_re(fname)
        target_keys = target_keys_for_region(own_key)
        if not target_keys:
            warnings.append({
                "유형": "지역 인식 실패", "시트": "-", "파일": fname,
                "설명": "파일명에서 지역명을 추출할 수 없습니다."
            })
            continue

        try:
            wb_src = openpyxl.load_workbook(fbytes, data_only=True)
        except Exception as e:
            warnings.append({"유형": "파일 읽기 오류", "시트": "-", "파일": fname, "설명": str(e)})
            continue

        for key in target_keys:
            target_sheet_name = JIJEOK_SHEET_MAP.get(key)
            if not target_sheet_name or target_sheet_name not in base_wb.sheetnames:
                warnings.append({
                    "유형": "시트 없음", "시트": target_sheet_name or "알수없음", "파일": fname,
                    "설명": f"총괄표 서식에 '{target_sheet_name}' 지역 시트가 없습니다."
                })
                continue

            if is_protected_summary_sheet(target_sheet_name):
                warnings.append({
                    "유형": "보호된 시트", "시트": target_sheet_name, "파일": fname,
                    "설명": "'총괄' 시트는 취합 대상에서 제외됩니다."
                })
                continue

            if target_sheet_name in wb_src.sheetnames:
                src_ws = wb_src[target_sheet_name]
            else:
                usable_sheets = [s for s in wb_src.worksheets if not is_protected_summary_sheet(s.title)]
                src_ws = usable_sheets[0] if usable_sheets else wb_src.worksheets[0]

            base_ws = base_wb[target_sheet_name]

            src_year_rows = {}
            for r in range(1, min(src_ws.max_row, 100) + 1):
                val = src_ws.cell(r, 1).value
                if val is not None:
                    val_str = str(val).replace(" ", "").strip()
                    if val_str == "합계" or val_str.endswith("년"):
                        src_year_rows[val_str] = r

            if not src_year_rows:
                warnings.append({
                    "유형": "데이터 없음", "시트": target_sheet_name, "파일": fname,
                    "설명": "원본에서 '합계'나 '연도' 데이터를 찾을 수 없습니다."
                })
                continue

            processed_count = 0
            max_col = min(base_ws.max_column, 30)

            for r in range(1, min(base_ws.max_row, 100) + 1):
                val = base_ws.cell(r, 1).value
                if val is not None:
                    val_str = str(val).replace(" ", "").strip()
                    if val_str in src_year_rows:
                        src_r = src_year_rows[val_str]
                        for c in range(2, max_col + 1):
                            base_cell = base_ws.cell(r, c)
                            if isinstance(base_cell, MergedCell) or is_formula(base_cell.value):
                                continue
                            src_val = src_ws.cell(src_r, c).value
                            base_cell.value = get_safe_value(src_val)
                        processed_count += 1

            if processed_count > 0:
                log.append({"파일": fname, "시트": target_sheet_name, "처리행수": processed_count})

    return base_wb, log, warnings
