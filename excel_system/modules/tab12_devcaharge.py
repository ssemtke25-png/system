"""
탭: 개발부담금 징수수수료 취합
(서식2) 총괄표 + 시군별 파일 → 4개 시트 취합

[시트별 취합 방식]
청구내역    : 자기 시군 행의 F(위임수수료), G(지급대상금액) 복사
월별납입내액 : 합계·4월·5월·6월 행의 B~K열 누적 합산
시군구별    : 자기 시군 행의 C~K열 복사 + 합계행 갱신
월별내역 3종: 자기 시군 행의 C~L열 복사 + 합계행 갱신

[특이사항]
- 가나다순(서식2 기준)
- 시군 파일이 자기 것만 남기고 서식 지워오는 경우 있음
- 수수료 없는 시군(모두 0)은 건너뜀
- xls 파일 자동 변환
"""
import io, re, zipfile, xml.etree.ElementTree as ET
import openpyxl
import streamlit as st
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

# 서식2 총괄표 시트명
BASE_SHEET_MAP = {
    "청구내역":    ["청구내역(총괄표-작성대상)", "청구내역(총괄표)", "청구내역"],
    "월별납입":    ["(월별납입내액 총괄표)"],
    "시군구별":    ["(시군구별 총괄표)"],
    "월4":         ["월별내역('26.4월-작성대상)", "월별내역('26.4월)"],
    "월5":         ["월별내역('26.5월-작성대상)", "월별내역('26.5월)"],
    "월6":         ["월별내역('26.6월-작성대상)", "월별내역('26.6월)"],
}

# 경북 시군 가나다순 (서식2 기준)
GYEONGBUK_ORDER = [
    "경산시","경주시","고령군","구미시","김천시","문경시","봉화군",
    "상주시","성주군","안동시","영덕군","영양군","영주시","영천시",
    "예천군","울릉군","울진군","의성군","청도군","청송군","칠곡군","포항시",
]

# ── 유틸 ─────────────────────────────────────────────────────────────
def shrink_styles(fb):
    fb.seek(0)
    try:
        zin = zipfile.ZipFile(fb, 'r')
        if 'xl/styles.xml' not in zin.namelist():
            fb.seek(0); return fb
        sb = zin.read('xl/styles.xml')
        ET.register_namespace('', NS_MAIN)
        root = ET.fromstring(sb)
        xfs = root.find(f'{{{NS_MAIN}}}cellStyleXfs')
        if xfs is None or len(xfs) <= 200:
            fb.seek(0); return fb
        first = xfs[0] if len(xfs) else None
        for x in list(xfs): xfs.remove(x)
        if first is not None: xfs.append(first)
        xfs.set('count','1')
        cxfs = root.find(f'{{{NS_MAIN}}}cellXfs')
        if cxfs:
            for x in cxfs:
                if x.get('xfId'): x.set('xfId','0')
        decl = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        nb = decl + ET.tostring(root, encoding='UTF-8')
        out = io.BytesIO()
        zout = zipfile.ZipFile(out,'w',zipfile.ZIP_DEFLATED)
        for item in zin.namelist():
            zout.writestr(item, nb if item=='xl/styles.xml' else zin.read(item))
        zout.close(); zin.close(); out.seek(0); return out
    except Exception:
        fb.seek(0); return fb


def load_wb(fb, data_only=True):
    return openpyxl.load_workbook(shrink_styles(fb), data_only=data_only)


def find_sheet(wb, candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    for name in wb.sheetnames:
        for c in candidates:
            if c.replace(' ','').replace("'","") in name.replace(' ','').replace("'",""):
                return wb[name]
    return None


def normalize_sigun(v):
    """셀값에서 시군명 정규화 (경주시, 경주 등 → 경주시)"""
    if not v: return None
    s = str(v).strip()
    # 경상북도 경주시 → 경주시
    s = re.sub(r'^경상북도\s*', '', s)
    s = re.sub(r'(시|군)$', '', s)
    for name in GYEONGBUK_ORDER:
        base = re.sub(r'(시|군)$','',name)
        if base in s:
            return name
    return None


def find_sigun_row(ws, sigun_name, search_col=2, start_row=6, end_row=35):
    """시트에서 해당 시군 행 번호 찾기"""
    base = re.sub(r'(시|군)$','',sigun_name)
    for r in range(start_row, min(end_row, ws.max_row+1)):
        v = ws.cell(r, search_col).value
        if v and base in str(v):
            return r
    return None


def find_sum_row(ws, search_col=2, start_row=5, end_row=12):
    """합계 행 찾기"""
    for r in range(start_row, min(end_row, ws.max_row+1)):
        v = ws.cell(r, search_col).value
        if v and '합계' in str(v):
            return r
        v1 = ws.cell(r, 1).value
        if v1 and '합계' in str(v1):
            return r
    return 7  # 기본값


def safe_num(v):
    if v is None or v == '' or v == '-': return 0
    try: return float(v)
    except: return 0


def has_data(wb_src):
    """수수료 데이터가 있는지 판단 (모두 0이면 False)"""
    ws = find_sheet(wb_src, BASE_SHEET_MAP["청구내역"])
    if not ws: return False
    for r in range(6, min(ws.max_row+1, 35)):
        v_fee = safe_num(ws.cell(r, 6).value)
        v_amt = safe_num(ws.cell(r, 7).value)
        if v_fee != 0 or v_amt != 0:
            return True
    return False


def extract_sigun_from_file(wb_src):
    """시군 파일에서 자기 시군명 추출"""
    # 1. 청구내역 시트에서 값이 있는 행의 시군명
    ws = find_sheet(wb_src, BASE_SHEET_MAP["청구내역"])
    if ws:
        for r in range(6, min(ws.max_row+1, 35)):
            v_fee = safe_num(ws.cell(r, 6).value)
            v_amt = safe_num(ws.cell(r, 7).value)
            if v_fee != 0 or v_amt != 0:
                name = normalize_sigun(ws.cell(r, 2).value)
                if name: return name
    # 2. 월별납입내액 기관명에서
    ws2 = find_sheet(wb_src, BASE_SHEET_MAP["월별납입"])
    if ws2:
        for r in range(1, 6):
            v = ws2.cell(r, 1).value
            if v and '기관명' in str(v):
                name = normalize_sigun(v)
                if name: return name
    return None


def extract_sigun_from_filename(fname):
    """파일명에서 시군명 추출"""
    m = re.match(r'^\d{1,3}[_.\s]+(.+?)\.', fname)
    raw = m.group(1) if m else fname
    raw = raw.replace('__','').replace('_',' ').strip()
    return normalize_sigun(raw)


# ── 시트별 취합 함수 ─────────────────────────────────────────────────

def fill_sheet1_청구내역(base_ws, src_wb, sigun_name, warns, fname):
    """청구내역: F열(위임수수료), G열(지급대상금액) 복사"""
    src_ws = find_sheet(src_wb, BASE_SHEET_MAP["청구내역"])
    if not src_ws: return 0

    # src에서 자기 시군 행 찾기
    src_row = find_sigun_row(src_ws, sigun_name)
    if not src_row:
        warns.append({"유형":"시군 행 없음","파일":fname,"설명":f"{sigun_name} 행을 src에서 못 찾음"})
        return 0

    # base에서 자기 시군 행 찾기
    base_row = find_sigun_row(base_ws, sigun_name)
    if not base_row:
        warns.append({"유형":"시군 행 없음","파일":fname,"설명":f"{sigun_name} 행을 base에서 못 찾음"})
        return 0

    copied = 0
    for col in [6, 7]:  # F, G열
        v = safe_num(src_ws.cell(src_row, col).value)
        cell = base_ws.cell(base_row, col)
        if not isinstance(cell, MergedCell):
            cell.value = safe_num(base_ws.cell(base_row, col).value) + v
            copied += 1
    return copied


def fill_sheet2_월별납입(base_ws, src_wb, warns, fname):
    """월별납입내액: 합계·4월·5월·6월 행 B~K열 누적합산"""
    src_ws = find_sheet(src_wb, BASE_SHEET_MAP["월별납입"])
    if not src_ws: return 0

    # 행 매핑: 합계=7, 4월=8, 5월=9, 6월=10
    DATA_ROWS = [7, 8, 9, 10]
    DATA_COLS = range(2, 12)  # B~K

    copied = 0
    for r in DATA_ROWS:
        for c in DATA_COLS:
            sv = safe_num(src_ws.cell(r, c).value)
            if sv == 0: continue
            base_cell = base_ws.cell(r, c)
            if not isinstance(base_cell, MergedCell):
                base_cell.value = safe_num(base_cell.value) + sv
                copied += 1
    return copied


def fill_sheet3_시군구별(base_ws, src_wb, sigun_name, warns, fname):
    """시군구별: 자기 시군 행 C~K열 복사 후 합계행 갱신"""
    src_ws = find_sheet(src_wb, BASE_SHEET_MAP["시군구별"])
    if not src_ws: return 0

    src_row = find_sigun_row(src_ws, sigun_name)
    if not src_row: return 0
    base_row = find_sigun_row(base_ws, sigun_name)
    if not base_row: return 0

    copied = 0
    for col in range(3, 13):  # C~L
        sv = safe_num(src_ws.cell(src_row, col).value)
        base_cell = base_ws.cell(base_row, col)
        if not isinstance(base_cell, MergedCell):
            base_cell.value = safe_num(base_cell.value) + sv
            copied += 1
    return copied


def fill_sheet4_월별내역(base_ws, src_wb, sigun_name, month_key, warns, fname):
    """월별내역(4·5·6월): 자기 시군 행 C~L열 복사"""
    src_ws = find_sheet(src_wb, BASE_SHEET_MAP[month_key])
    if not src_ws: return 0

    src_row = find_sigun_row(src_ws, sigun_name)
    if not src_row: return 0
    base_row = find_sigun_row(base_ws, sigun_name)
    if not base_row: return 0

    copied = 0
    for col in range(3, 13):  # C~L
        sv = safe_num(src_ws.cell(src_row, col).value)
        base_cell = base_ws.cell(base_row, col)
        if not isinstance(base_cell, MergedCell):
            base_cell.value = safe_num(base_cell.value) + sv
            copied += 1
    return copied


def update_sum_rows(base_wb):
    """모든 시트의 합계행 갱신"""
    # 청구내역 합계행(행6) F,G열
    ws1 = find_sheet(base_wb, BASE_SHEET_MAP["청구내역"])
    if ws1:
        sum_row = find_sum_row(ws1, search_col=2, start_row=5, end_row=8)
        for col in [6, 7]:
            total = sum(
                safe_num(ws1.cell(r, col).value)
                for r in range(sum_row+1, ws1.max_row+1)
                if ws1.cell(r, 1).value is not None or ws1.cell(r, 2).value is not None
            )
            c = ws1.cell(sum_row, col)
            if not isinstance(c, MergedCell): c.value = total

    # 시군구별 합계행(행7) C~L열
    ws3 = find_sheet(base_wb, BASE_SHEET_MAP["시군구별"])
    if ws3:
        sum_row = find_sum_row(ws3, search_col=2, start_row=5, end_row=9)
        for col in range(3, 13):
            total = sum(
                safe_num(ws3.cell(r, col).value)
                for r in range(sum_row+1, ws3.max_row+1)
                if ws3.cell(r, 2).value is not None
            )
            c = ws3.cell(sum_row, col)
            if not isinstance(c, MergedCell): c.value = total

    # 월별내역 각 시트 합계행(행7) C~L열
    for mk in ["월4","월5","월6"]:
        ws = find_sheet(base_wb, BASE_SHEET_MAP[mk])
        if not ws: continue
        sum_row = find_sum_row(ws, search_col=1, start_row=5, end_row=9)
        for col in range(3, 13):
            total = sum(
                safe_num(ws.cell(r, col).value)
                for r in range(sum_row+1, ws.max_row+1)
                if ws.cell(r, 2).value is not None
            )
            c = ws.cell(sum_row, col)
            if not isinstance(c, MergedCell): c.value = total


# ── 메인 취합 ────────────────────────────────────────────────────────
def aggregate(template_bytes, region_files):
    log, warns = [], []

    base_wb = load_wb(io.BytesIO(template_bytes), data_only=False)

    # base 시트 확인
    base_sheets = {
        "청구내역": find_sheet(base_wb, BASE_SHEET_MAP["청구내역"]),
        "월별납입": find_sheet(base_wb, BASE_SHEET_MAP["월별납입"]),
        "시군구별": find_sheet(base_wb, BASE_SHEET_MAP["시군구별"]),
        "월4":       find_sheet(base_wb, BASE_SHEET_MAP["월4"]),
        "월5":       find_sheet(base_wb, BASE_SHEET_MAP["월5"]),
        "월6":       find_sheet(base_wb, BASE_SHEET_MAP["월6"]),
    }
    missing = [k for k,v in base_sheets.items() if not v]
    if missing:
        raise ValueError(f"총괄표에서 시트를 찾지 못했습니다: {missing}")

    # 시군 파일 정렬 (파일명 앞 번호 기준)
    def sort_key(item):
        m = re.match(r'^(\d{1,3})', item[0])
        return int(m.group(1)) if m else 999

    sorted_files = sorted(region_files, key=sort_key)

    skipped, processed = [], []

    for fname, fbytes in sorted_files:
        try:
            src_wb = load_wb(io.BytesIO(fbytes), data_only=True)
        except Exception as e:
            warns.append({"유형":"파일 오류","파일":fname,"설명":str(e)})
            continue

        # 수수료 없으면 건너뜀
        if not has_data(src_wb):
            skipped.append(fname)
            continue

        # 시군명 파악
        sigun = extract_sigun_from_file(src_wb) or extract_sigun_from_filename(fname)
        if not sigun:
            warns.append({"유형":"시군 인식 실패","파일":fname,"설명":"시군명을 파악하지 못함"})
            continue

        n1 = fill_sheet1_청구내역(base_sheets["청구내역"], src_wb, sigun, warns, fname)
        n2 = fill_sheet2_월별납입(base_sheets["월별납입"], src_wb, warns, fname)
        n3 = fill_sheet3_시군구별(base_sheets["시군구별"], src_wb, sigun, warns, fname)
        n4 = fill_sheet4_월별내역(base_sheets["월4"], src_wb, sigun, "월4", warns, fname)
        n5 = fill_sheet4_월별내역(base_sheets["월5"], src_wb, sigun, "월5", warns, fname)
        n6 = fill_sheet4_월별내역(base_sheets["월6"], src_wb, sigun, "월6", warns, fname)

        processed.append(sigun)
        log.append({"파일":fname,"시군":sigun,"청구내역":n1,"월별납입":n2,"시군구별":n3,"4월":n4,"5월":n5,"6월":n6})

    # 합계행 갱신
    update_sum_rows(base_wb)

    return base_wb, log, warns, processed, skipped


# ── Streamlit UI ─────────────────────────────────────────────────────
def render():
    st.caption("시군별 개발부담금 징수수수료 파일을 받아 총괄표(서식2)에 취합합니다.")
    st.info(
        "📌 수수료가 없는 시군(전체 0)은 자동으로 건너뜁니다.\n\n"
        "📌 서식을 지우고 자기 것만 남긴 파일도 자동 처리됩니다.\n\n"
        "📌 색이 채워진 칸(계산용)은 절대 덮어쓰지 않습니다."
    )

    tpl_file = st.file_uploader(
        "① 총괄표(서식2) 파일 업로드",
        type=["xlsx"],
        key="tpl_gaebul"
    )
    region_files = st.file_uploader(
        "② 시군별 파일 업로드 (여러 개, xls/xlsx 모두 가능)",
        type=["xlsx","xls"],
        accept_multiple_files=True,
        key="region_gaebul"
    )

    if tpl_file and region_files and st.button("🚀 취합 시작", key="btn_gaebul"):
        # xls → xlsx 변환
        converted = []
        for f in region_files:
            raw = f.read()
            if f.name.lower().endswith('.xls'):
                try:
                    import xlrd
                    book = xlrd.open_workbook(file_contents=raw)
                    wb_new = openpyxl.Workbook()
                    wb_new.remove(wb_new.active)
                    for sname in book.sheet_names():
                        ws_s = book.sheet_by_name(sname)
                        ws_d = wb_new.create_sheet(sname)
                        for r in range(ws_s.nrows):
                            for c in range(ws_s.ncols):
                                ws_d.cell(r+1, c+1, ws_s.cell_value(r, c))
                    buf = io.BytesIO()
                    wb_new.save(buf)
                    converted.append((f.name, buf.getvalue()))
                except Exception as e:
                    st.warning(f"⚠️ {f.name} xls 변환 실패: {e}")
                    converted.append((f.name, raw))
            else:
                converted.append((f.name, raw))

        try:
            result_wb, log, warns, processed, skipped = aggregate(
                tpl_file.read(), converted
            )

            out = io.BytesIO()
            result_wb.save(out)

            st.success(f"✅ 취합 완료! 처리 {len(processed)}개 시군")

            if skipped:
                st.info(f"📋 수수료 없어 건너뜀: {', '.join([s.split('/')[-1] for s in skipped])}")

            st.download_button(
                "📥 취합 결과 다운로드",
                data=out.getvalue(),
                file_name="개발부담금_취합결과.xlsx",
                key="dl_gaebul"
            )

            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**처리된 시군 (가나다순)**")
                for s in processed:
                    st.write(f"  ✅ {s}")
            with col2:
                if warns:
                    st.warning(f"⚠️ 확인 필요 {len(warns)}건")
                    st.dataframe(warns, use_container_width=True)

            with st.expander("처리 로그"):
                st.dataframe(log, use_container_width=True)

        except ValueError as e:
            st.error(f"❌ {e}")
        except Exception as e:
            st.error(f"오류: {e}")
            st.exception(e)
