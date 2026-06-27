"""
탭3: 실거래 월보
시군구별 '실거래 과태료' 파일을 받아 하나로 취합한다.

1·2번과 구조가 다름:
 - "총괄표" 시트: 'N. ...' 제목으로 여러 개의 하위 표가 들어있고, 표마다
   지역별 블록 크기가 다름(예: 9행/6행/1행). 표별로 독립적으로 자기 지역
   구간을 찾아 그 구간만 src에서 가져와 채움. (포항은 남/북 2개 구간 모두 처리)
 - 그 외 시트(과태료 처분 세부내역 / 세무관서 통보내역 / 불법거래신고 처리현황):
   개별 신고 건이 한 줄씩 쌓이는 원시 데이터 시트라서, 합산이 아니라
   시군 순서대로 아래로 이어붙임(연결).
 - 셀 색칠(테마 색상)이 있으면 "건드리면 안 되는 계산 칸"이라는 표시이므로
   수식과 마찬가지로 절대 덮어쓰지 않음. 단 theme=0(흰색)은 보호 대상에서 제외.
 - 원본 파일들의 styles.xml에 쓰이지 않는 스타일 정의가 수만 건씩 남아있어
   로딩이 매우 느려지는 경우가 있어, 읽기 전 스타일을 자동으로 정리해 속도를 높임.
 - "내역없음" 같은 안내문이 셀 분리 입력된 행, 합계 수식의 좁은 범위,
   서식 누락 등 실제 운영 데이터에서 발견된 여러 비정상 패턴을 자동 보정.
"""
import io
import re
import zipfile
from copy import copy as _copy_style
import xml.etree.ElementTree as ET

import openpyxl
import streamlit as st
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from modules.common import (
    is_formula, region_key, VALID_REGION_KEYS, target_keys_for_region,
)

VALID_REGION_KEYS_RE = VALID_REGION_KEYS
REGION_ORDER_RE = [
    '포항남', '포항북', '경주', '김천', '안동', '구미', '영주', '영천', '상주', '문경',
    '경산', '군위', '의성', '청송', '영양', '영덕', '청도', '고령', '성주', '칠곡',
    '예천', '봉화', '울진', '울릉'
]

TOTAL_SHEET_NAME = '총괄표'
APPEND_SHEET_NAMES = ['과태료 처분 세부내역', '세무관서 통보내역', '불법거래신고 처리현황']

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def is_valid_region_re(label):
    return region_key(label) in VALID_REGION_KEYS_RE or region_key(label) == '군위'


def extract_own_region_re(filename):
    m = re.match(r'^\d{0,3}[_.\s]*([가-힣]+)', filename)
    if m:
        key = region_key(m.group(1))
        if key:
            return key
    for k in REGION_ORDER_RE:
        if k and k in filename:
            return k
    return None


def has_protective_color(cell):
    """셀에 색(테마 또는 RGB)이 채워져 있으면 '건드리면 안 되는 칸'으로 간주.
    단, theme=0은 흰색(배경색)이라 실제로는 '색 없음'과 같으므로 보호 대상에서 제외."""
    fill = cell.fill
    if not getattr(fill, 'patternType', None):
        return False
    fg = getattr(fill, 'fgColor', None)
    if not fg:
        return False

    fg_type = getattr(fg, 'type', None)
    if fg_type == 'theme':
        theme = getattr(fg, 'theme', None)
        if theme is not None and theme != 0:
            return True
    elif fg_type == 'rgb':
        rgb = getattr(fg, 'rgb', None)
        if rgb not in (None, '00000000', 'FFFFFFFF'):
            return True
    return False


def is_protected_re(cell):
    return is_formula(cell.value) or has_protective_color(cell)


def shrink_styles_xml(file_bytes):
    """
    xlsx 파일(BytesIO)에서 xl/styles.xml의 cellStyleXfs/cellStyles를 최소화해서 파일을 가볍게 만든다.
    cellXfs(실제 셀이 직접 참조하는 서식 - 글꼴/배경색/테두리)는 그대로 보존하므로
    화면에 보이는 서식은 바뀌지 않는다. xfId(이름 붙은 스타일 참조)만 기본값(0)으로 통일한다.
    어떤 이유로든 처리에 실패하면 원본을 그대로 반환한다(안전 우선).
    """
    file_bytes.seek(0)
    try:
        zin = zipfile.ZipFile(file_bytes, 'r')
        names = zin.namelist()

        if 'xl/styles.xml' not in names:
            file_bytes.seek(0)
            return file_bytes

        styles_bytes = zin.read('xl/styles.xml')

        ET.register_namespace('', NS_MAIN)
        root = ET.fromstring(styles_bytes)

        cell_style_xfs = root.find(f'{{{NS_MAIN}}}cellStyleXfs')
        cell_xfs = root.find(f'{{{NS_MAIN}}}cellXfs')
        cell_styles = root.find(f'{{{NS_MAIN}}}cellStyles')

        if cell_style_xfs is None or cell_xfs is None:
            file_bytes.seek(0)
            return file_bytes

        old_count = len(cell_style_xfs)
        if old_count <= 200:
            file_bytes.seek(0)
            return file_bytes

        first_xf = cell_style_xfs[0] if len(cell_style_xfs) > 0 else None
        for xf in list(cell_style_xfs):
            cell_style_xfs.remove(xf)
        if first_xf is not None:
            cell_style_xfs.append(first_xf)
        cell_style_xfs.set('count', '1')

        if cell_styles is not None:
            normal_style = None
            for cs in cell_styles:
                if cs.get('builtinId') == '0' or cs.get('name') in ('Normal', '표준'):
                    normal_style = cs
                    break
            if normal_style is None and len(cell_styles) > 0:
                normal_style = cell_styles[0]
            for cs in list(cell_styles):
                cell_styles.remove(cs)
            if normal_style is not None:
                normal_style.set('xfId', '0')
                cell_styles.append(normal_style)
                cell_styles.set('count', '1')

        for xf in cell_xfs:
            if xf.get('xfId') is not None:
                xf.set('xfId', '0')

        xml_decl = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        new_styles_bytes = xml_decl + ET.tostring(root, encoding='UTF-8')

        out_buffer = io.BytesIO()
        zout = zipfile.ZipFile(out_buffer, 'w', zipfile.ZIP_DEFLATED)
        for item in names:
            if item == 'xl/styles.xml':
                zout.writestr(item, new_styles_bytes)
            else:
                zout.writestr(item, zin.read(item))
        zout.close()
        zin.close()

        out_buffer.seek(0)
        return out_buffer
    except Exception:
        file_bytes.seek(0)
        return file_bytes


def normalize_excessive_merges(wb, max_col_limit=60):
    """가로로 비정상적으로 넓은(예: A2:XFD2처럼 16384열까지) 병합 영역을 안전한 크기로 축소.
    이런 극단적으로 넓은 머지는 일부 환경에서 파일 손상을 유발할 수 있어,
    실제 쓰이는 범위만 남기고 정리한다. 셀 내용(좌상단 값)은 그대로 보존."""
    for ws in wb.worksheets:
        to_fix = []
        for mc in list(ws.merged_cells.ranges):
            if (mc.max_col - mc.min_col) > max_col_limit:
                to_fix.append(mc)
        for mc in to_fix:
            min_row, min_col, max_row = mc.min_row, mc.min_col, mc.max_row
            ws.unmerge_cells(str(mc))
            new_max_col = min(min_col + max_col_limit, mc.max_col)
            if new_max_col > min_col:
                ws.merge_cells(start_row=min_row, start_column=min_col,
                                end_row=max_row, end_column=new_max_col)


def find_table_boundaries(ws, max_scan_row=500):
    """'N. ...' 형식의 제목이 있는 행들을 찾아 표 구간을 나눔."""
    title_rows = []
    max_row = min(ws.max_row, max_scan_row)
    for r in range(1, max_row + 1):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and re.match(r'^\d+\.\s', v.strip()):
            title_rows.append(r)
    boundaries = []
    for i, start in enumerate(title_rows):
        end = title_rows[i + 1] - 1 if i + 1 < len(title_rows) else max_row
        boundaries.append((start, end))
    return boundaries


def find_region_blocks_in_range(ws, start_row, end_row):
    blocks = []
    for r in range(start_row, end_row + 1):
        label = ws.cell(r, 1).value
        if is_valid_region_re(label):
            blocks.append((region_key(label) or '군위', r))
    return blocks


def get_block_size_re(blocks, end_row):
    if len(blocks) >= 2:
        return blocks[1][1] - blocks[0][1]
    elif len(blocks) == 1:
        return max(1, end_row - blocks[0][1] + 1)
    return 1


def fill_total_sheet_re(base_ws, src_ws, own_key, warnings, sheet_title, fname):
    """총괄표 시트: 제목으로 나뉜 표마다, 자기 지역 구간을 src에서 찾아 base의 같은 구간으로 복사.
    포항처럼 한 파일이 2개 구간(포항남/포항북)을 모두 채워야 하는 경우도 처리한다."""
    base_tables = find_table_boundaries(base_ws)
    src_tables = find_table_boundaries(src_ws)

    total_count = 0
    n_tables = min(len(base_tables), len(src_tables))

    if len(base_tables) != len(src_tables):
        warnings.append({
            "유형": "표 개수 다름", "시트": sheet_title, "파일": fname,
            "설명": f"총괄표는 표 {len(base_tables)}개, 이 파일은 표 {len(src_tables)}개 - 앞쪽 {n_tables}개만 매칭"
        })

    target_keys = target_keys_for_region(own_key)

    for t in range(n_tables):
        b_start, b_end = base_tables[t]
        s_start, s_end = src_tables[t]

        base_blocks = find_region_blocks_in_range(base_ws, b_start, b_end)
        src_blocks = find_region_blocks_in_range(src_ws, s_start, s_end)

        base_row_for_key = dict(base_blocks)
        src_row_for_key = dict(src_blocks)

        block_size = get_block_size_re(base_blocks, b_end)
        max_col = max(base_ws.max_column, 5)

        for key in target_keys:
            base_start_row = base_row_for_key.get(key)
            src_start_row = src_row_for_key.get(key)

            if base_start_row is None or src_start_row is None:
                continue

            for offset in range(block_size):
                gr_base = base_start_row + offset
                gr_src = src_start_row + offset
                if gr_base > b_end:
                    break
                for c in range(1, max_col + 1):
                    base_cell = base_ws.cell(gr_base, c)
                    if isinstance(base_cell, MergedCell):
                        continue
                    if is_protected_re(base_cell):
                        continue
                    src_val = src_ws.cell(gr_src, c).value
                    if src_val is None:
                        continue
                    if isinstance(src_val, str) and src_val.startswith('#'):
                        warnings.append({
                            "유형": "오류 값 발견", "시트": sheet_title,
                            "셀": f"{get_column_letter(c)}{gr_src}", "파일": fname,
                            "설명": f"원본 값이 '{src_val}'(수식 오류)이라 이 칸은 건너뜀. 누계 합산은 계속 진행됨"
                        })
                        continue
                    base_cell.value = src_val
                total_count += 1

    return total_count


def looks_like_note_text(v):
    """'* ', 'ㅇ', '※' 등으로 시작하는 안내문 텍스트인지 판별 (데이터가 아님)."""
    if not isinstance(v, str):
        return False
    s = v.strip()
    return s.startswith(('*', 'ㅇ', '※', '<'))


def find_data_start_row_re(ws, max_scan_row=10):
    """'합계' 텍스트가 있는 행의 다음 행을 데이터 시작행으로 판단."""
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        for c in range(1, 4):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and v.strip() == '합계':
                return r + 1
    return 4


def find_data_end_row_re(ws, start_row, max_scan_row=3000):
    """데이터 시작행부터, A~E열 전체가 비어있는 행이 연속 4개 나오거나
    A열이 안내문 텍스트인 행을 만나면 그 직전까지를 데이터로 간주."""
    max_row = min(ws.max_row, max_scan_row)
    empty_streak = 0
    last_data_row = start_row - 1
    for r in range(start_row, max_row + 1):
        a_val = ws.cell(r, 1).value
        if looks_like_note_text(a_val):
            break
        row_has_data = any(ws.cell(r, c).value is not None for c in range(1, 6))
        if row_has_data:
            last_data_row = r
            empty_streak = 0
        else:
            empty_streak += 1
            if empty_streak >= 4:
                break
    return last_data_row


def normalize_data_area_style(ws, start_row, end_row, max_col_limit=40):
    """서식 파일을 만들 때 예시 데이터 몇 줄에만 서식(정렬, 줄바꿈 등)을 입혀두고
    그 아래는 서식이 비어있는 경우가 있다. 이러면 실제 데이터가 늘어났을 때
    행마다 표시가 들쭥날쭥해 보인다. 데이터 영역의 첫 행 서식을 기준으로
    삼아, 서식이 비어있는 나머지 행에도 똑같이 입혀서 표 전체가 일관되게 보이도록 한다.
    글자 값은 절대 건드리지 않고 서식(글꼴/정렬/테두리/배경색)만 복사한다."""
    if end_row < start_row:
        return
    max_col = min(ws.max_column, max_col_limit)

    template_cells = {}
    for c in range(1, max_col + 1):
        cell = ws.cell(start_row, c)
        if not isinstance(cell, MergedCell):
            template_cells[c] = cell

    for r in range(start_row + 1, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            if cell.alignment is not None and cell.alignment.horizontal is not None:
                continue
            tmpl = template_cells.get(c)
            if tmpl is None:
                continue
            cell.font = _copy_style(tmpl.font)
            cell.alignment = _copy_style(tmpl.alignment)
            cell.border = _copy_style(tmpl.border)
            cell.number_format = tmpl.number_format
            if not has_protective_color(cell):
                cell.fill = _copy_style(tmpl.fill)


def unmerge_in_data_area(ws, start_row, max_scan_row=3000):
    """데이터가 채워질 영역(start_row 이후)에 남아있는 병합 셀을 모두 해제한다.
    서식 작성 시 예시 데이터 몇 줄에만 맞춰 만든 머지가, 실제 데이터가 그보다
    많아지면 데이터 영역 한가운데를 가로막아 일부 칸이 누락되는 문제를 막기 위함.
    헤더 영역(start_row 이전)의 머지는 건드리지 않는다."""
    max_row = min(ws.max_row, max_scan_row)
    to_unmerge = []
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= start_row and mc.min_row <= max_row:
            to_unmerge.append(str(mc))
    for ref in to_unmerge:
        ws.unmerge_cells(ref)


def clear_existing_data_area(ws, start_row, max_scan_row=3000, max_col_limit=40):
    """이어붙이기 전, base 시트에 이미 들어있던 예시 데이터나 안내문 잔재를 미리 비움."""
    max_row = min(ws.max_row, max_scan_row)
    max_col = min(ws.max_column, max_col_limit)
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def looks_like_split_note_row(ws, r, max_col=10):
    """'내역없음', '해당없음'처럼 안내 문구가 셀 병합 없이 한 글자씩 여러 칸에
    나뉘어 입력된 행인지 감지한다. (예: C='내', D='역', E='없', F='음')
    조건: 연속된 칸 2개 이상이 각각 정확히 한 글자(공백 제외)인 경우."""
    single_char_cols = []
    for c in range(1, min(max_col, 10) + 1):
        v = ws.cell(r, c).value
        if isinstance(v, str) and len(v.strip()) == 1:
            single_char_cols.append(c)
    if len(single_char_cols) < 2:
        return False
    return all(single_char_cols[i + 1] - single_char_cols[i] == 1 for i in range(len(single_char_cols) - 1))


def is_valid_real_row(ws, r, max_col):
    """이 줄이 '진짜 취합해야 할 유효한 데이터'인지 스마트하게 판별합니다."""
    text_concat = ""
    for c in range(1, min(max_col, 10) + 1):
        v = ws.cell(r, c).value
        if v and isinstance(v, str):
            text_concat += v.replace(" ", "")

    if any(x in text_concat for x in ["해당없음", "해당사항없음", "실적없음", "내역없음", "내용없음"]):
        return False

    if looks_like_split_note_row(ws, r, max_col):
        return False

    has_real_data = False
    for c in range(6, max_col + 1):
        v = ws.cell(r, c).value
        if v is not None and str(v).strip() not in ['', '-', '0']:
            has_real_data = True
            break

    return has_real_data


def append_sheet_data(base_ws, src_ws, base_next_row, warnings, sheet_title, fname, max_col_limit=40):
    """src 시트의 데이터 중 '진짜 실적'만 골라내어 base_ws에 빈틈없이 이어붙임.
    '과태료 처분 세부내역'은 3행 1블록 구조라 블록 단위로 유효성을 판단한다."""
    src_start = find_data_start_row_re(src_ws)
    src_end = find_data_end_row_re(src_ws, src_start)

    if src_end < src_start:
        return base_next_row, 0

    max_col = min(max(src_ws.max_column, base_ws.max_column), max_col_limit)
    added_rows = 0

    if sheet_title == '과태료 처분 세부내역':
        sr = src_start
        while sr <= src_end:
            is_block_valid = False
            for offset in range(3):
                if sr + offset <= src_end and is_valid_real_row(src_ws, sr + offset, max_col):
                    is_block_valid = True
                    break

            if is_block_valid:
                br = base_next_row + added_rows
                for offset in range(3):
                    curr_sr = sr + offset
                    curr_br = br + offset
                    if curr_sr > src_end:
                        break
                    for c in range(1, max_col + 1):
                        base_cell = base_ws.cell(curr_br, c)
                        if isinstance(base_cell, MergedCell):
                            continue
                        src_val = src_ws.cell(curr_sr, c).value
                        if isinstance(src_val, str) and src_val.startswith('#'):
                            warnings.append({
                                "유형": "오류 값 발견", "시트": sheet_title,
                                "셀": f"{get_column_letter(c)}{curr_sr}", "파일": fname,
                                "설명": f"원본 값이 '{src_val}'(수식 오류)이라 빈 값으로 처리. 나머지 데이터는 그대로 이어붙임"
                            })
                            src_val = None
                        base_cell.value = src_val
                added_rows += 3
            sr += 3
    else:
        for sr in range(src_start, src_end + 1):
            if not is_valid_real_row(src_ws, sr, max_col):
                continue

            br = base_next_row + added_rows
            for c in range(1, max_col + 1):
                base_cell = base_ws.cell(br, c)
                if isinstance(base_cell, MergedCell):
                    continue
                src_val = src_ws.cell(sr, c).value
                if isinstance(src_val, str) and src_val.startswith('#'):
                    warnings.append({
                        "유형": "오류 값 발견", "시트": sheet_title,
                        "셀": f"{get_column_letter(c)}{sr}", "파일": fname,
                        "설명": f"원본 값이 '{src_val}'(수식 오류)이라 빈 값으로 처리. 나머지 데이터는 그대로 이어붙임"
                    })
                    src_val = None
                base_cell.value = src_val
            added_rows += 1

    return base_next_row + added_rows, added_rows


def expand_count_formula_range(formula, new_end_row):
    """'=COUNTA(B4:B5)'처럼 범위 기반 합계 수식을 만나면, 그 범위의 끝 행을
    new_end_row까지 자동으로 넓혀준다. (수식 자체를 절대 값으로 바꾸지 않고,
    범위만 실제 데이터 끝까지 확장 - 작성금지 칸 보호 원칙은 그대로 유지)
    함수명(COUNTA/SUM/COUNT 등)과 무관하게 동작하며, 이미 범위가 충분히 넓으면 그대로 둔다."""
    m = re.match(r'^=(\w+)\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', formula)
    if not m:
        return formula, False
    func, col1, row1, col2, row2 = m.groups()
    row1, row2 = int(row1), int(row2)
    if new_end_row <= row2:
        return formula, False
    return f"={func}({col1}{row1}:{col2}{new_end_row})", True


def update_total_count_re(base_ws):
    """'합계' 행을 찾아 그 옆 칸을 정확하게 갱신한다.
    - 그 칸이 좁은 범위의 합계 수식(예: =COUNTA(B4:B5))이면, 범위를 실제 데이터
      끝까지 넓혀서 수식 자체는 보존한 채로 정확한 값이 나오게 한다.
    - 수식이 아니라 일반 값이면, 실제 데이터 건수로 직접 갱신한다."""
    for r in range(1, min(base_ws.max_row, 10) + 1):
        for c in range(1, 4):
            v = base_ws.cell(r, c).value
            if v and isinstance(v, str) and v.strip() == '합계':
                cnt_cell = base_ws.cell(r, c + 1)
                start = r + 1

                if start > base_ws.max_row:
                    if not is_formula(cnt_cell.value):
                        cnt_cell.value = 0
                    return

                actual_rows = 0
                last_data_row = start - 1
                for rr in range(start, base_ws.max_row + 1):
                    if any(base_ws.cell(rr, cc).value is not None for cc in range(1, base_ws.max_column + 1)):
                        actual_rows += 1
                        last_data_row = rr

                if is_formula(cnt_cell.value):
                    new_formula, changed = expand_count_formula_range(cnt_cell.value, last_data_row)
                    if changed:
                        cnt_cell.value = new_formula
                    return

                if base_ws.title == '과태료 처분 세부내역':
                    cnt_cell.value = actual_rows // 3
                else:
                    cnt_cell.value = actual_rows
                return


def sort_key_for_filename_re(fname):
    """파일명 앞의 번호(01,02...)로 정렬. 번호가 없으면 지역 순서로."""
    m = re.match(r'^(\d{1,3})[_.\s]', fname)
    if m:
        return int(m.group(1))
    own_key = extract_own_region_re(fname)
    if own_key in REGION_ORDER_RE:
        return 100 + REGION_ORDER_RE.index(own_key)
    return 999


def fill_real_estate_template(template_bytes, region_files_with_names):
    """
    template_bytes: 총괄표(서식) 파일
    region_files_with_names: [(파일명, BytesIO), ...]
    """
    log = []
    warnings = []

    template_bytes = shrink_styles_xml(template_bytes)
    base_wb = openpyxl.load_workbook(template_bytes, data_only=False)
    normalize_excessive_merges(base_wb)

    sorted_files = sorted(region_files_with_names, key=lambda x: sort_key_for_filename_re(x[0]))

    # 각 시군 파일을 한 번만 로딩해서 재사용 (스타일 축소 후 로딩 -> 속도 개선)
    loaded_wbs = {}
    for fname, fbytes in sorted_files:
        try:
            shrunk = shrink_styles_xml(fbytes)
            wb_src = openpyxl.load_workbook(shrunk, data_only=True)
            loaded_wbs[fname] = wb_src
        except Exception as e:
            warnings.append({
                "유형": "파일 읽기 오류", "시트": "-", "파일": fname,
                "설명": f"파일을 여는 중 오류가 발생하여 건너뜀: {e}"
            })

    # --- 총괄표 시트 채우기 ---
    if TOTAL_SHEET_NAME in base_wb.sheetnames:
        base_total_ws = base_wb[TOTAL_SHEET_NAME]
        for fname, fbytes in sorted_files:
            if fname not in loaded_wbs:
                continue
            own_key = extract_own_region_re(fname)
            if not own_key:
                warnings.append({
                    "유형": "지역 인식 실패", "시트": TOTAL_SHEET_NAME, "파일": fname,
                    "설명": "파일명에서 지역명을 추출할 수 없어 건너뜀"
                })
                continue
            wb_src = loaded_wbs[fname]
            if TOTAL_SHEET_NAME not in wb_src.sheetnames:
                warnings.append({
                    "유형": "시트 없음", "시트": TOTAL_SHEET_NAME, "파일": fname,
                    "설명": f"'{TOTAL_SHEET_NAME}' 시트가 없어 건너뜀"
                })
                continue
            src_ws = wb_src[TOTAL_SHEET_NAME]
            n = fill_total_sheet_re(base_total_ws, src_ws, own_key, warnings, TOTAL_SHEET_NAME, fname)
            if n:
                log.append({"파일": fname, "시트": TOTAL_SHEET_NAME, "처리행수": n})

    # --- 이어붙이기 시트들 처리 ---
    for sheet_name in APPEND_SHEET_NAMES:
        if sheet_name not in base_wb.sheetnames:
            continue
        base_ws = base_wb[sheet_name]
        next_row = find_data_start_row_re(base_ws)
        unmerge_in_data_area(base_ws, next_row)
        clear_existing_data_area(base_ws, next_row)

        for fname, fbytes in sorted_files:
            if fname not in loaded_wbs:
                continue
            wb_src = loaded_wbs[fname]
            if sheet_name not in wb_src.sheetnames:
                warnings.append({
                    "유형": "시트 없음", "시트": sheet_name, "파일": fname,
                    "설명": f"'{sheet_name}' 시트가 없어 건너뜀"
                })
                continue
            src_ws = wb_src[sheet_name]
            next_row, n_rows = append_sheet_data(base_ws, src_ws, next_row, warnings, sheet_name, fname)
            if n_rows:
                log.append({"파일": fname, "시트": sheet_name, "처리행수": n_rows})

        data_start_row = find_data_start_row_re(base_ws)
        if next_row - 1 >= data_start_row:
            normalize_data_area_style(base_ws, data_start_row, next_row - 1)

        update_total_count_re(base_ws)

    return base_wb, log, warnings


def render():
    """탭3 화면을 그린다. app.py에서 with tab3: render() 형태로 호출."""
    st.caption(
        "시군구별 '실거래 월보(과태료)' 파일을 받아 하나로 취합합니다. "
        "'총괄표' 시트는 자기 지역 구간만 찾아 채우고, 나머지 시트(세부내역 등)는 시군 순서대로 이어붙입니다."
    )
    st.info(
        "📌 셀에 색이 채워진 칸(계산용 칸)은 수식과 마찬가지로 절대 덮어쓰지 않습니다.\n\n"
        "📌 일부 시군 파일이 누락되어도 나머지 파일은 정상 처리되며, 총괄표(서식) 자체는 영향받지 않습니다.\n\n"
        "📌 원본 값이 수식 오류(#VALUE! 등)인 경우, 해당 칸만 비우고 화면에 오류로 표시합니다. 전체 취합(누계)은 계속 진행됩니다."
    )

    template_file3 = st.file_uploader("① 총괄표(서식) 파일 업로드", type=["xlsx"], key="template_up3")
    region_files3 = st.file_uploader("② 시군구별 파일 업로드 (여러 개)", type=["xlsx"], accept_multiple_files=True, key="region_up3")

    if template_file3 and region_files3 and st.button("🚀 취합 시작", key="btn3"):
        try:
            template_bytes3 = io.BytesIO(template_file3.read())
            region_files_with_names3 = [(f.name, io.BytesIO(f.read())) for f in region_files3]

            result_wb3, log3, warns3 = fill_real_estate_template(template_bytes3, region_files_with_names3)

            o3 = io.BytesIO()
            result_wb3.save(o3)

            st.success("취합이 완료되었습니다.")
            st.download_button("📥 다운로드", o3.getvalue(), "실거래_월보_결과.xlsx", key="dl3")

            if warns3:
                st.warning(f"⚠️ 확인이 필요한 항목 {len(warns3)}건이 발견되었습니다. (취합 결과는 정상 생성됨)")
                st.dataframe(warns3, use_container_width=True)
            else:
                st.info("특이사항 없이 정상적으로 취합되었습니다.")

            with st.expander("처리 로그 보기 (성공 내역)"):
                st.dataframe(log3, use_container_width=True)

        except Exception as e:
            st.error(f"오류: {e}")
            st.exception(e)
