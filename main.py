import os
import sys
import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
import webbrowser
import tempfile
import subprocess
import calendar
from datetime import datetime
import json
import re
import traceback

# ==========================================
# [0. 경로 설정 - data 폴더 완전 분리]
# ==========================================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

EXCEL_PATH = os.path.join(DATA_DIR, "data.xlsx")
EVENT_PATH = os.path.join(DATA_DIR, "calendar_events.json")
LAW_HTML_PATH = os.path.join(DATA_DIR, "지적재조사에 관한 특별법(인용조문 3단비교).html")
REG_DIR = os.path.join(DATA_DIR, "규정")
MANUAL_NAME = "질의회신_프로그램_매뉴얼.hwpx"

# ==========================================
# [1. 윈도우 생성 및 GUI 기본 설정]
# ==========================================
root = tk.Tk()
root.title("지적재조사 통합 업무지원 시스템 v1.0(경상북도 지적재조사팀)")
root.geometry("950x980")
root.configure(bg="#f4f6f7")

style = ttk.Style()
style.theme_use("clam")
style.configure("TLabelframe", background="#f4f6f7", bordercolor="#bdc3c7", borderwidth=1)
style.configure("TLabelframe.Label", font=("Malgun Gothic", 11, "bold"), foreground="#2c3e50", background="#f4f6f7")

# ==========================================
# [2. 데이터 파싱 및 🌟탐정 모드 에러 알림🌟]
# ==========================================
def parse_law_data(file_path):
    if not os.path.exists(file_path): 
        messagebox.showwarning("파일 누락", f"법령 파일을 찾을 수 없습니다.\ndata 폴더를 확인하세요.")
        return []
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            soup = BeautifulSoup(f, "html.parser")
        law_db = []
        for tr in soup.select("tr"):
            tds = tr.select("td")
            if len(tds) >= 3 and (jo_span := tds[0].select_one("span.bl")):
                law_db.append({
                    "조문": jo_span.get_text(strip=True), 
                    "법률": tds[0].get_text("\n", strip=True), 
                    "시행령": tds[1].get_text("\n", strip=True), 
                    "시행규칙": tds[2].get_text("\n", strip=True)
                })
        return law_db
    except Exception as e: 
        messagebox.showerror("법령 로드 치명적 오류", f"법령 데이터를 읽는 중 오류가 발생했습니다.\n\n{traceback.format_exc()}")
        return []

def parse_reg_data(file_path):
    if not os.path.exists(file_path): return []
    try:
        with open(file_path, "r", encoding="utf-8") as f: html_content = f.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, "r", encoding="cp949") as f: html_content = f.read()
        except: return []
    except: return []

    soup = BeautifulSoup(html_content, "html.parser")
    reg_list = []
    
    for tr in soup.select("tr"):
        tds = tr.select("td")
        if tds and (jo_span := tr.select_one("span.bl")):
            reg_list.append({"조문": jo_span.get_text(strip=True), "내용": "\n".join([td.get_text("\n", strip=True) for td in tds])})
    if reg_list: return reg_list
    
    text_lines = soup.get_text(separator="\n").split("\n")
    current_jo, current_content = "전체 내용", []
    for line in text_lines:
        line = line.strip()
        if not line: continue
        if re.match(r'^제\s*\d+\s*조', line) or re.match(r'^\[별표', line):
            if current_content: reg_list.append({"조문": current_jo, "내용": "\n".join(current_content)})
            current_jo, current_content = line, [line]
        else: current_content.append(line)
    if current_content: reg_list.append({"조문": current_jo, "내용": "\n".join(current_content)})
    return reg_list

law_db = parse_law_data(LAW_HTML_PATH)
if not os.path.exists(REG_DIR):
    messagebox.showwarning("폴더 누락", f"'규정' 폴더가 data 폴더 안에 없습니다.\n규정 검색 기능이 제한됩니다.")
    reg_db = {}
else:
    reg_db = {f.replace(".html", ""): parse_reg_data(os.path.join(REG_DIR, f)) for f in os.listdir(REG_DIR) if f.endswith(".html")}

def safe_load_sheet(path, sheet_name, columns):
    if not os.path.exists(path):
        messagebox.showerror("파일 누락", f"엑셀 파일이 없습니다!\ndata 폴더에 data.xlsx를 넣어주세요.")
        return pd.DataFrame(columns=columns)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        actual_sheet = None
        for sn in wb.sheetnames:
            if sn.replace(" ", "") == sheet_name.replace(" ", ""):
                actual_sheet = sn
                break
                
        if not actual_sheet:
            messagebox.showwarning("시트 누락", f"엑셀 파일에 '{sheet_name}' 시트가 없습니다.\n현재 존재하는 시트: {wb.sheetnames}")
            return pd.DataFrame(columns=columns)
            
        rows = [list(r) for r in wb[actual_sheet].iter_rows(min_row=2, values_only=True) if any(r)]
        df = pd.DataFrame(rows)
        if df.empty:
            messagebox.showwarning("데이터 없음", f"'{actual_sheet}' 시트는 존재하지만 데이터가 없습니다. 2행부터 내용이 있는지 확인하세요.")
            return pd.DataFrame(columns=columns)
            
        for i in range(len(df.columns), len(columns)): df[i] = ""
        df = df.iloc[:, :len(columns)]
        df.columns = columns
        return df.fillna("").astype(str)
    except PermissionError:
        messagebox.showerror("접근 거부", "엑셀 파일이 현재 열려있습니다!\n엑셀 창을 닫고 프로그램을 다시 실행해주세요.")
        return pd.DataFrame(columns=columns)
    except Exception as e: 
        messagebox.showerror("엑셀 로드 치명적 오류", f"엑셀 파일을 읽지 못했습니다.\n\n상세원인:\n{traceback.format_exc()}")
        return pd.DataFrame(columns=columns)

df_qna = safe_load_sheet(EXCEL_PATH, "질의회신", ["제목", "내용", "수정여부"])
df_case = safe_load_sheet(EXCEL_PATH, "판례검색", ["제목", "내용", "수정여부"])
current_results = {"질의회신": df_qna.copy(), "판례검색": df_case.copy()}

def load_events():
    if os.path.exists(EVENT_PATH):
        try:
            with open(EVENT_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
                return {k: ({"memo": v, "use_alarm": False, "alarm_days": 1} if isinstance(v, str) else v) for k, v in data.items()}
        except: return {}
    return {}

# ==========================================
# [3. 검색 및 화면 렌더링 로직]
# ==========================================
def open_law_popup_window(jo_text):
    clean_search = jo_text.replace(" ", "")
    target = None
    for item in law_db:
        clean_jo = item['조문'].replace(" ", "")
        if clean_jo == clean_search or clean_jo.startswith(clean_search + "("):
            target = item
            break
    if not target:
        messagebox.showinfo("알림", f"[{jo_text}]에 해당하는 법령 정보를 찾을 수 없습니다.")
        return
    popup = tk.Toplevel(root)
    popup.title(f"법령 상세 - {target['조문']}")
    popup.geometry("600x500")
    txt = tk.Text(popup, font=("Malgun Gothic", 11), padx=15, pady=15, bg="#fdfdfd")
    txt.pack(fill=tk.BOTH, expand=True)
    content = f"📜 [법률]\n{target['법률']}\n\n{'='*40}\n⚙️ [시행령]\n{target['시행령']}\n\n{'='*40}\n📝 [시행규칙]\n{target['시행규칙']}"
    txt.insert(tk.END, content)
    txt.config(state=tk.DISABLED)

def on_law_link_click(event):
    idx = detail_text.index(f"@{event.x},{event.y}")
    ranges = detail_text.tag_ranges("law_link")
    for i in range(0, len(ranges), 2):
        if detail_text.compare(ranges[i], "<=", idx) and detail_text.compare(idx, "<=", ranges[i+1]):
            clicked_jo = detail_text.get(ranges[i], ranges[i+1])
            open_law_popup_window(clicked_jo)
            break

def search_data(*args):
    try:
        global current_results
        mode, keyword = mode_var.get(), search_entry.get().strip()
        result_listbox.delete(0, tk.END)
        
        if mode in ["질의회신", "판례검색"]:
            target_df = df_qna if mode == "질의회신" else df_case
            if not keyword: current_results[mode] = target_df.copy()
            else:
                if only_title_var.get() == 1: current_results[mode] = target_df[target_df['제목'].str.contains(keyword, case=False, na=False)]
                else: current_results[mode] = target_df[target_df['제목'].str.contains(keyword, case=False, na=False) | target_df['내용'].str.contains(keyword, case=False, na=False)]
            
            icon = "📑" if mode == "질의회신" else "⚖️"
            for idx, row in current_results[mode].iterrows():
                result_listbox.insert(tk.END, f" {icon} {row['제목']}")
                if str(row.get("수정여부")).strip().upper() == "Y":
                    result_listbox.itemconfig(result_listbox.size() - 1, fg="#27ae60", bg="#e8f8f5")
        else:
            for item in law_db:
                if not keyword or (only_title_var.get() == 1 and keyword in item['조문']) or (only_title_var.get() == 0 and any(keyword in item[k] for k in ['조문', '법률', '시행령', '시행규칙'])):
                    result_listbox.insert(tk.END, f" [법령] {item['조문']}")
            for reg_name, reg_data in reg_db.items():
                display_name = reg_name.replace("규정_", "")
                for item in reg_data:
                    if not keyword or (only_title_var.get() == 1 and keyword in item['조문']) or (only_title_var.get() == 0 and (keyword in item['조문'] or keyword in item['내용'])):
                        result_listbox.insert(tk.END, f" [{display_name}] {item['조문']}")
    except Exception as e:
        messagebox.showerror("검색 화면 오류", f"화면에 데이터를 표시하는 중 오류가 발생했습니다.\n\n{traceback.format_exc()}")

def show_detail(event):
    if not result_listbox.curselection(): return
    selected_text = result_listbox.get(result_listbox.curselection()[0])
    keyword = search_entry.get().strip()
    current_mode = mode_var.get() 
    
    detail_text.config(state=tk.NORMAL)
    detail_text.delete("1.0", tk.END)
    
    if current_mode == "질의회신": 
        detail_text.insert(tk.END, str(df_qna.loc[current_results["질의회신"].index[result_listbox.curselection()[0]], '내용']))
    elif current_mode == "판례검색": 
        detail_text.insert(tk.END, str(df_case.loc[current_results["판례검색"].index[result_listbox.curselection()[0]], '내용']))
    elif "[법령]" in selected_text:
        jo_title = selected_text.split("] ")[1].strip()
        target = next((item for item in law_db if item['조문'] == jo_title), None)
        if target: detail_text.insert(tk.END, f"⚖️ {target['조문']} 3단 비교 정보\n{'='*60}\n\n📜 [법률]\n{target['법률']}\n\n⚙️ [시행령]\n{target['시행령']}\n\n📝 [시행규칙]\n{target['시행규칙']}\n")
    elif "[" in selected_text and "]" in selected_text:
        d_name, j_title = selected_text.split("]")[0].split("[")[1].strip(), selected_text.split("] ")[1].strip()
        for r_name, r_data in reg_db.items():
            if d_name in r_name:
                target = next((item for item in r_data if item['조문'] == j_title), None)
                if target: detail_text.insert(tk.END, f"📖 [{d_name}] {target['조문']}\n{'='*60}\n\n{target['내용']}\n")
                break

    if keyword:
        start_pos = "1.0"
        while True:
            start_pos = detail_text.search(keyword, start_pos, stopindex=tk.END, nocase=True)
            if not start_pos: break
            end_pos = f"{start_pos}+{len(keyword)}c"
            detail_text.tag_add("highlight", start_pos, end_pos)
            start_pos = end_pos

    if current_mode == "질의회신":
        content = detail_text.get("1.0", tk.END)
        for match in re.finditer(r"제\s*\d+\s*조(?:의\s*\d+)?", content):
            start_idx, end_idx = f"1.0 + {match.start()} chars", f"1.0 + {match.end()} chars"
            detail_text.tag_add("law_link", start_idx, end_idx)
    else:
        detail_text.tag_remove("law_link", "1.0", tk.END)
        
    detail_text.config(state=tk.DISABLED)

def clear_search(): search_entry.delete(0, tk.END); search_data()

# ==========================================
# [4. 스마트 일정 플래너]
# ==========================================
def open_calendar_popup():
    cal_win = tk.Toplevel(root)
    cal_win.title("지적재조사팀 일정 관리 플래너 v4.0")
    cal_win.geometry("580x780")
    cal_win.configure(bg="#f8f9fa")
    
    now = datetime.now()
    cur_year, cur_month = tk.IntVar(value=now.year), tk.IntVar(value=now.month)
    selected_date_str = tk.StringVar(value=f"{now.year}-{now.month:02d}-{now.day:02d}")
    
    header_frame = tk.Frame(cal_win, bg="#2c3e50", padx=10, pady=10)
    header_frame.pack(fill=tk.X)
    tk.Label(header_frame, text="📅 팀 일정 플래너 (D-Day 카운팅 지원)", font=("Malgun Gothic", 12, "bold"), fg="white", bg="#2c3e50").pack(side=tk.LEFT)
    
    control_frame = tk.Frame(cal_win, bg="#f8f9fa", pady=10)
    control_frame.pack(fill=tk.X)
    cal_frame = tk.Frame(cal_win, bg="white", bd=1, relief="solid")
    cal_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
    
    memo_group = ttk.LabelFrame(cal_win, text=" 📝 선택된 날짜 일정 입력 및 알람 설정 ", padding=10)
    memo_group.pack(fill=tk.X, padx=15, pady=15)
    
    option_line = tk.Frame(memo_group, bg="#f4f6f7", pady=5)
    option_line.pack(fill=tk.X)
    
    alarm_var = tk.BooleanVar(value=False)
    tk.Checkbutton(option_line, text="🔔 알람 켜기", variable=alarm_var, bg="#f4f6f7", font=("Malgun Gothic", 10, "bold"), fg="#e67e22").pack(side=tk.LEFT, padx=(0, 10))
    date_label = tk.Label(option_line, text=f"선택일: {selected_date_str.get()}", font=("Malgun Gothic", 10, "bold"), fg="#2980b9", bg="#f4f6f7")
    date_label.pack(side=tk.LEFT, padx=5)
    
    tk.Label(option_line, text="알람기간:", bg="#f4f6f7", font=("Malgun Gothic", 9)).pack(side=tk.LEFT, padx=(15, 2))
    
    days_map = {"당일": 0, "1일 전": 1, "2일 전": 2, "3일 전": 3, "5일 전": 5, "7일 전": 7, "10일 전": 10, "20일 전": 20, "30일 전": 30}
    alarm_days_var = tk.StringVar(value="1일 전")
    ttk.Combobox(option_line, textvariable=alarm_days_var, values=list(days_map.keys()), width=8, state="readonly").pack(side=tk.LEFT)
    
    event_text_area = tk.Text(memo_group, font=("Malgun Gothic", 11), bd=1, relief="solid", height=4)
    event_text_area.pack(fill=tk.X, pady=(5, 5))
    
    def save_event():
        events, d_key = load_events(), selected_date_str.get()
        memo, is_alarm = event_text_area.get("1.0", tk.END).strip(), alarm_var.get()
        if memo: events[d_key] = {"memo": memo, "use_alarm": is_alarm, "alarm_days": days_map.get(alarm_days_var.get(), 1)}
        elif d_key in events: del events[d_key]
        try:
            with open(EVENT_PATH, "w", encoding="utf-8") as f: json.dump(events, f, ensure_ascii=False, indent=4)
            messagebox.showinfo("저장 성공", f"[{d_key}] 일정이 저장되었습니다.")
            update_main_banner(); draw_calendar()
        except Exception as e: messagebox.showerror("오류", f"저장 오류: {e}")

    tk.Button(memo_group, text="💾 일정저장", bg="#2ecc71", fg="white", font=("Malgun Gothic", 10, "bold"), relief="flat", padx=15, command=save_event).pack(side=tk.RIGHT, pady=(0, 5))

    def click_date(day):
        if day == 0: return
        d_key = f"{cur_year.get()}-{cur_month.get():02d}-{day:02d}"
        selected_date_str.set(d_key); date_label.config(text=f"선택일: {d_key}")
        event_text_area.delete("1.0", tk.END)
        events = load_events()
        if d_key in events:
            event_text_area.insert("1.0", events[d_key]["memo"])
            alarm_var.set(events[d_key].get("use_alarm", False))
            days_rev_map = {v: k for k, v in days_map.items()}
            alarm_days_var.set(days_rev_map.get(events[d_key].get("alarm_days", 1), "1일 전"))
        else:
            alarm_var.set(False); alarm_days_var.set("1일 전")
        draw_calendar()
            
    def draw_calendar():
        for w in cal_frame.winfo_children(): w.destroy()
        events = load_events()
        for i, (d, col) in enumerate(zip(["일", "월", "화", "수", "목", "금", "토"], ["#e74c3c", "#333", "#333", "#333", "#333", "#333", "#3498db"])):
            tk.Label(cal_frame, text=d, font=("Malgun Gothic", 10, "bold"), fg=col, bg="#eee", pady=5).grid(row=0, column=i, sticky="nsew")
        
        month_days = calendar.Calendar(firstweekday=6).monthdayscalendar(cur_year.get(), cur_month.get())
        for r_idx, week in enumerate(month_days):
            for c_idx, day in enumerate(week):
                if day == 0: tk.Label(cal_frame, bg="white").grid(row=r_idx+1, column=c_idx, sticky="nsew")
                else:
                    d_key = f"{cur_year.get()}-{cur_month.get():02d}-{day:02d}"
                    has_ev, is_al = d_key in events, events.get(d_key, {}).get("use_alarm", False)
                    lines = [l[:6]+".." if len(l)>6 else l for l in events[d_key]["memo"].split('\n') if l.strip()] if has_ev else []
                    
                    d_txt = f"{day} 🔔\n" if is_al else f"{day}\n"
                    d_txt += "\n".join(lines[:2]) + ("\n외 .." if len(lines)>2 else "\n" if not lines else "")
                    
                    bg_col = "#e8f4f8" if selected_date_str.get() == d_key else ("#fffaf0" if is_al else "#f4fbf7" if has_ev else "white")
                    fg_col = "#e67e22" if is_al else ("#27ae60" if has_ev and c_idx not in (0,6) else ["#e74c3c", "#333", "#333", "#333", "#333", "#333", "#3498db"][c_idx])
                    
                    tk.Button(cal_frame, text=d_txt, font=("Malgun Gothic", 9, "bold" if has_ev else "normal"), fg=fg_col, bg=bg_col, relief="flat", activebackground="#d5dbdb", justify="center", anchor="n", pady=4, command=lambda d=day: click_date(d)).grid(row=r_idx+1, column=c_idx, sticky="nsew", padx=1, pady=1)
                    
        for i in range(7): cal_frame.columnconfigure(i, weight=1)
        for i in range(len(month_days)+1): cal_frame.rowconfigure(i, weight=1)
        month_label.config(text=f"{cur_year.get()}년 {cur_month.get()}월")

    tk.Button(control_frame, text="◀", font=("Arial", 11), relief="flat", command=lambda: [cur_year.set(cur_year.get()-1) if cur_month.get()==1 else None, cur_month.set(12 if cur_month.get()==1 else cur_month.get()-1), draw_calendar()], bg="#f8f9fa").pack(side=tk.LEFT, padx=20)
    month_label = tk.Label(control_frame, font=("Malgun Gothic", 12, "bold"), bg="#f8f9fa"); month_label.pack(side=tk.LEFT, expand=True)
    tk.Button(control_frame, text="▶", font=("Arial", 11), relief="flat", command=lambda: [cur_year.set(cur_year.get()+1) if cur_month.get()==12 else None, cur_month.set(1 if cur_month.get()==12 else cur_month.get()+1), draw_calendar()], bg="#f8f9fa").pack(side=tk.RIGHT, padx=20)
    draw_calendar(); click_date(now.day)

def get_upcoming_alarms():
    events, upcoming = load_events(), []
    for d_str, info in events.items():
        if not info.get("use_alarm", False): continue
        try:
            delta = (datetime.strptime(d_str, "%Y-%m-%d").date() - datetime.now().date()).days
            if 0 <= delta <= info.get("alarm_days", 1): 
                upcoming.append({"date": d_str, "d_day": delta, "memo": info["memo"]})
        except: continue
    return sorted(upcoming, key=lambda x: x["d_day"])

# ==========================================
# [5. 보조 기능]
# ==========================================
def print_content():
    if not (content := detail_text.get("1.0", tk.END).strip()): return
    html = f"<!DOCTYPE html><html><head><meta charset='utf-8'><style>@page {{ size: A4; margin: 15mm; }} body {{ font-family: 'Malgun Gothic'; font-size: 13pt; line-height: 1.4; }} .box {{ border: 1.5px solid #000; padding: 15px; }} .hd {{ font-weight: bold; border-bottom: 1px solid #000; padding-bottom: 6px; margin-bottom: 12px; }}</style></head><body><div class='box'><div class='hd'>■ 지적재조사팀 출력물</div><div>{content.replace(chr(10), '<br>')}</div></div></body></html>"
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8") as f: f.write(html)
        subprocess.run(["powershell", "-Command", f'Start-Process "rundll32.exe" -ArgumentList "mshtml.dll,PrintHTML `"{f.name}`"" -WindowStyle Normal'], shell=True)
    except Exception as e: messagebox.showerror("오류", f"인쇄 오류: {e}")

def open_notepad():
    try: os.system("start notepad.exe")
    except: pass

def open_manual():
    target_path = os.path.join(DATA_DIR, MANUAL_NAME)
    if os.path.exists(target_path):
        try: os.startfile(target_path)
        except Exception as e: messagebox.showerror("오류", f"매뉴얼을 여는 중 오류 발생: {e}")
    else:
        messagebox.showwarning("파일 없음", f"매뉴얼 파일({MANUAL_NAME})이 없습니다.\ndata 폴더에 넣어주세요.")

# ==========================================
# [6. GUI 화면 레이아웃 (배너 완벽 복원!)]
# ==========================================
# [수정 1] 배너 박스를 만들고, 그 안에 들어갈 '글씨'를 꼭 화면에 붙여넣습니다! (.pack 추가)
banner_frame = tk.Frame(root, bg="#fff3cd", bd=1, relief="solid")
banner_label = tk.Label(banner_frame, bg="#fff3cd", pady=6, font=("Malgun Gothic", 11, "bold"), fg="#e74c3c")
banner_label.pack() 

current_font_size = 11

search_frame = tk.Frame(root, bg="#2c3e50", padx=15, pady=15)
search_frame.pack(fill=tk.X, padx=10, pady=(5, 5))
tk.Label(search_frame, text="🔍 통합 검색", font=("Malgun Gothic", 12, "bold"), fg="white", bg="#2c3e50").pack(side=tk.LEFT, padx=(5, 15))

# [수정 2] 일정이 있을 때 배너를 띄우는 함수 (항상 검색창 '위에' 뜨도록 고정!)
def update_main_banner():
    upcoming = get_upcoming_alarms()
    if upcoming:
        first = upcoming[0]
        d_text = "[오늘]" if first["d_day"] == 0 else f"[{first['d_day']}일 후]"
        banner_label.config(text=f"🔔 중요 예정 알림: {d_text} {(first['memo'].replace(chr(10), ' / ')[:40] + '...') if len(first['memo'])>40 else first['memo']}")
        banner_frame.pack(before=search_frame, fill=tk.X, padx=10, pady=(10, 0)) # 무조건 검색창(search_frame) 위에 착! 붙게 설정
    else:
        banner_frame.pack_forget()

entry_container = tk.Frame(search_frame, bg="white", borderwidth=0)
entry_container.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=4)

mode_var = tk.StringVar(value="질의회신")
mode_select = ttk.Combobox(entry_container, textvariable=mode_var, values=["질의회신", "법령검색","판례검색"], width=10, state="readonly")
mode_select.pack(side=tk.LEFT, padx=(5, 5)); mode_select.bind("<<ComboboxSelected>>", search_data)

search_entry = tk.Entry(entry_container, font=("Malgun Gothic", 12), borderwidth=0, highlightthickness=0)
search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5); search_entry.bind("<KeyRelease>", search_data)
tk.Button(search_frame, text="✕", command=clear_search, bg="#e74c3c", fg="white", relief="flat", font=("Arial", 11, "bold"), padx=15).pack(side=tk.LEFT, padx=(5, 15))

only_title_var = tk.IntVar(value=1)
tk.Checkbutton(search_frame, text="제목만 검색", variable=only_title_var, command=search_data, bg="#2c3e50", fg="white", selectcolor="#2c3e50").pack(side=tk.LEFT, padx=(0, 15))
tk.Button(search_frame, text="검 색", command=search_data, bg="#3498db", fg="white", padx=30, font=("Malgun Gothic", 12, "bold"), relief="flat").pack(side=tk.LEFT)

list_title_frame = tk.Frame(root, bg="#f4f6f7")
tk.Label(list_title_frame, text=" 검색 결과 목록 ", font=("Malgun Gothic", 11, "bold"), fg="#2c3e50", bg="#f4f6f7").pack(side=tk.LEFT)
tk.Button(list_title_frame, text="📖 매뉴얼", bg="#3498db", fg="white", font=("Malgun Gothic", 9, "bold"), relief="flat", padx=12, pady=1, command=open_manual).pack(side=tk.RIGHT, padx=5)

list_group = ttk.LabelFrame(root, labelwidget=list_title_frame, padding=10)
list_group.pack(fill=tk.BOTH, expand=True, padx=15, pady=5)
list_scrollbar = tk.Scrollbar(list_group); list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
result_listbox = tk.Listbox(list_group, font=("Malgun Gothic", 11), yscrollcommand=list_scrollbar.set, borderwidth=1, relief="solid", highlightthickness=0, selectbackground="#3498db")
result_listbox.pack(fill=tk.BOTH, expand=True, padx=5, pady=5); result_listbox.bind("<<ListboxSelect>>", show_detail)
list_scrollbar.config(command=result_listbox.yview)

detail_group = ttk.LabelFrame(root, text=" 상세 내용 보기 ", padding=10); detail_group.pack(fill=tk.BOTH, expand=True, padx=15, pady=(5, 15))
btn_bar = tk.Frame(detail_group, bg="#f4f6f7", pady=2); btn_bar.pack(fill=tk.X, padx=5)

tk.Label(btn_bar, text="글자크기:", bg="#f4f6f7", font=("Malgun Gothic", 10), fg="#7f8c8d").pack(side=tk.LEFT, padx=(0, 2))
tk.Button(btn_bar, text="가+", command=lambda: [globals().update(current_font_size=min(24, current_font_size+2)), detail_text.config(font=("Malgun Gothic", current_font_size))], bg="#e0e0e0", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8).pack(side=tk.LEFT, padx=2)
tk.Button(btn_bar, text="가-", command=lambda: [globals().update(current_font_size=max(8, current_font_size-2)), detail_text.config(font=("Malgun Gothic", current_font_size))], bg="#e0e0e0", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8).pack(side=tk.LEFT, padx=2)

tk.Button(btn_bar, text="프린트", bg="#e67e22", fg="white", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8, command=print_content).pack(side=tk.LEFT, padx=4)
tk.Button(btn_bar, text="법제처", bg="#2c3e50", fg="white", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8, command=lambda: webbrowser.open("https://www.law.go.kr")).pack(side=tk.LEFT, padx=4)
tk.Button(btn_bar, text="바른땅", bg="#2980b9", fg="white", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8, command=lambda: webbrowser.open("https://work.newjijuk.go.kr")).pack(side=tk.LEFT, padx=4)
tk.Button(btn_bar, text="메모장", bg="#7f8c8d", fg="white", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8, command=lambda: os.system("start notepad.exe")).pack(side=tk.LEFT, padx=4)
tk.Button(btn_bar, text="일정", bg="#9b59b6", fg="white", relief="flat", font=("Malgun Gothic", 9, "bold"), padx=8, command=open_calendar_popup).pack(side=tk.LEFT, padx=4)

text_container = tk.Frame(detail_group, bd=1, relief="solid")
text_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=(10, 5))

detail_scrollbar = tk.Scrollbar(text_container)
detail_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

detail_text = tk.Text(text_container, font=("Malgun Gothic", 11), fg="#1a5276", borderwidth=0, highlightthickness=0, padx=15, pady=15, spacing1=4, yscrollcommand=detail_scrollbar.set, state=tk.DISABLED)
detail_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
detail_scrollbar.config(command=detail_text.yview)

detail_text.tag_configure("highlight", background="yellow", foreground="black")
detail_text.tag_configure("law_link", foreground="#e74c3c", underline=True)
detail_text.tag_bind("law_link", "<Enter>", lambda e: detail_text.config(cursor="hand2"))
detail_text.tag_bind("law_link", "<Leave>", lambda e: detail_text.config(cursor=""))
detail_text.tag_bind("law_link", "<Button-1>", on_law_link_click)

# ==========================================
# [7. 시작 시 점검 보고서 및 알림 출력]
# ==========================================
def show_diagnostic_report():
    msg = f"""📊 시스템 자체 점검 보고서 📊

✅ 질의회신: {len(df_qna)} 건 로드됨
✅ 판례검색: {len(df_case)} 건 로드됨
✅ 법령조문: {len(law_db)} 건 로드됨
✅ 규정문서: {len(reg_db)} 개 로드됨

※ 만약 '0건'인 항목이 있다면, 엑셀 시트 이름이나 파일 위치를 확인하세요!"""
    messagebox.showinfo("데이터 로드 결과", msg)

def check_welcome_alarm():
    if not (upcoming := get_upcoming_alarms()): return
    alarm_msg = "주무관님! 중요한 알람 일정이 예정되어 있습니다.\n" + "═"*45 + "\n"
    for item in upcoming:
        d_text = "★ [오늘] ★" if item["d_day"] == 0 else f"▶ [{item['d_day']}일 후 (D-{item['d_day']})]"
        alarm_msg += f"{d_text} ({item['date']})\n   : {item['memo'][:30]}...\n\n"
    messagebox.showinfo("📅 중요 업무 리마인더", alarm_msg.strip())

search_data()
update_main_banner() # 프로그램 켜질 때 상단 배너를 검색창 위로 착! 올려줍니다.
root.after(500, show_diagnostic_report) 
root.after(1000, check_welcome_alarm)
root.mainloop()